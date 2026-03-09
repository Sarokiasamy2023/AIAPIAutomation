from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks, Form, Depends
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import create_engine, Column, Integer, String, DateTime, Boolean, Float, ForeignKey, Text
from sqlalchemy.orm import sessionmaker, relationship, Session, declarative_base
from pydantic import BaseModel, ConfigDict
from typing import List, Optional
from datetime import datetime
import os
import pandas as pd
from openai import AzureOpenAI
from pathlib import Path
from io import BytesIO
from dotenv import load_dotenv
import requests
import json
import httpx
from schema_validator import SchemaValidator

load_dotenv()

app = FastAPI(title="GS API Test Platform")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

DATABASE_URL = "sqlite:///./gs_api_test.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

class Mapping(Base):
    __tablename__ = "mappings"
    
    id = Column(Integer, primary_key=True, index=True)
    filename = Column(String, nullable=False)
    upload_date = Column(DateTime, default=datetime.utcnow)
    status = Column(String, default="uploaded")
    parsed_count = Column(Integer, default=0)
    business_rules = Column(Text)
    selected_endpoint = Column(String)
    
    parsed_fields = relationship("ParsedField", back_populates="mapping", cascade="all, delete-orphan")
    test_scenarios = relationship("TestScenario", back_populates="mapping", cascade="all, delete-orphan")

class ParsedField(Base):
    __tablename__ = "parsed_fields"
    
    id = Column(Integer, primary_key=True, index=True)
    mapping_id = Column(Integer, ForeignKey("mappings.id"), nullable=False)
    field_name = Column(String, nullable=False)
    data_type = Column(String)
    required = Column(Boolean, default=False)
    example_value = Column(String)
    gs_rule = Column(Text)
    expected_response = Column(Text)
    rule_type = Column(String)
    confidence = Column(Float)
    routing = Column(String)
    approved = Column(Boolean, default=False)
    
    mapping = relationship("Mapping", back_populates="parsed_fields")

class TestScenario(Base):
    __tablename__ = "test_scenarios"
    
    id = Column(Integer, primary_key=True, index=True)
    mapping_id = Column(Integer, ForeignKey("mappings.id"), nullable=True)
    endpoint_id = Column(Integer, ForeignKey("api_endpoints.id"))
    name = Column(String, nullable=False)
    description = Column(Text)
    category = Column(String)
    status = Column(String, default="pending")
    request_body = Column(Text)
    expected_response = Column(Text)
    json_schema = Column(Text)
    business_rule_id = Column(Integer, ForeignKey("custom_business_rules.id"), nullable=True)
    
    mapping = relationship("Mapping", back_populates="test_scenarios")
    endpoint = relationship("APIEndpoint", back_populates="test_scenarios")
    validation_results = relationship("ValidationResult", back_populates="scenario", cascade="all, delete-orphan")
    executions = relationship("TestExecution", back_populates="scenario", cascade="all, delete-orphan")

class TestExecution(Base):
    __tablename__ = "test_executions"
    
    id = Column(Integer, primary_key=True, index=True)
    scenario_id = Column(Integer, ForeignKey("test_scenarios.id"), nullable=False)
    execution_date = Column(DateTime, default=datetime.utcnow)
    status = Column(String)
    pass_count = Column(Integer, default=0)
    fail_count = Column(Integer, default=0)
    total_response_time_ms = Column(Integer)
    request_body = Column(Text)
    response_body = Column(Text)
    expected_response = Column(Text)
    
    scenario = relationship("TestScenario", back_populates="executions")
    validation_results = relationship("ValidationResult", back_populates="execution", cascade="all, delete-orphan")

class ValidationResult(Base):
    __tablename__ = "validation_results"
    
    id = Column(Integer, primary_key=True, index=True)
    scenario_id = Column(Integer, ForeignKey("test_scenarios.id"), nullable=False)
    execution_id = Column(Integer, ForeignKey("test_executions.id"))
    field_name = Column(String, nullable=False)
    expected = Column(String)
    actual = Column(String)
    status = Column(String)
    root_cause = Column(Text)
    root_cause_category = Column(String)
    suggested_action = Column(Text)
    ai_confidence = Column(Integer)
    validation_type = Column(String)
    response_time_ms = Column(Integer)
    status_code = Column(Integer)
    timestamp = Column(DateTime, default=datetime.utcnow)
    
    scenario = relationship("TestScenario", back_populates="validation_results")
    execution = relationship("TestExecution", back_populates="validation_results")

class RootCauseCache(Base):
    __tablename__ = "root_cause_cache"
    
    id = Column(Integer, primary_key=True, index=True)
    cache_key = Column(String, unique=True, index=True)
    category = Column(String)
    explanation = Column(Text)
    suggested_action = Column(Text)
    confidence = Column(Integer)
    created_at = Column(DateTime, default=datetime.utcnow)

class Environment(Base):
    __tablename__ = "environments"
    
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False, unique=True)
    description = Column(String)
    variables = Column(Text)
    is_active = Column(Boolean, default=False)
    created_date = Column(DateTime, default=datetime.utcnow)
    
    endpoints = relationship("APIEndpoint", back_populates="environment", cascade="all, delete-orphan")

class APIEndpoint(Base):
    __tablename__ = "api_endpoints"
    
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    environment_id = Column(Integer, ForeignKey("environments.id"), nullable=False)
    base_url = Column(String, nullable=False)
    method = Column(String, default="GET")
    path = Column(String, default="/")
    auth_type = Column(String)
    auth_token = Column(String)
    headers = Column(Text)
    default_request_body = Column(Text)
    timeout_ms = Column(Integer, default=5000)
    expected_status = Column(Integer, default=200)
    max_response_time_ms = Column(Integer, default=2000)
    created_date = Column(DateTime, default=datetime.utcnow)
    token_endpoint_id = Column(Integer, ForeignKey("api_endpoints.id"))
    token_response_path = Column(String, default="access_token")
    
    environment = relationship("Environment", back_populates="endpoints")
    test_scenarios = relationship("TestScenario", back_populates="endpoint")

class TokenCache(Base):
    __tablename__ = "token_cache"
    
    id = Column(Integer, primary_key=True, index=True)
    endpoint_id = Column(Integer, ForeignKey("api_endpoints.id"), nullable=False)
    token = Column(String, nullable=False)
    expires_at = Column(DateTime, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)

class CustomBusinessRule(Base):
    __tablename__ = "custom_business_rules"
    
    id = Column(Integer, primary_key=True, index=True)
    endpoint_id = Column(Integer, ForeignKey("api_endpoints.id"), nullable=False)
    rule_name = Column(String, nullable=False)
    rule_text = Column(Text, nullable=False)
    created_date = Column(DateTime, default=datetime.utcnow)
    updated_date = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    is_active = Column(Boolean, default=True)

class BusinessRuleValidation(Base):
    __tablename__ = "business_rule_validations"
    
    id = Column(Integer, primary_key=True, index=True)
    rule_id = Column(Integer, ForeignKey("custom_business_rules.id"), nullable=False)
    endpoint_id = Column(Integer, ForeignKey("api_endpoints.id"), nullable=False)
    execution_date = Column(DateTime, default=datetime.utcnow)
    record_number = Column(Integer)
    field_name = Column(String)
    expected_value = Column(String)
    actual_value = Column(String)
    result = Column(String)
    error_message = Column(Text)
    rule_text = Column(Text)

Base.metadata.create_all(bind=engine)

# Migration: Add business_rule_id column and fix mapping_id constraint
try:
    from sqlalchemy import inspect, text
    inspector = inspect(engine)
    columns = [col['name'] for col in inspector.get_columns('test_scenarios')]
    
    # Add business_rule_id column if it doesn't exist
    if 'business_rule_id' not in columns:
        print("Adding business_rule_id column to test_scenarios table...")
        with engine.connect() as conn:
            conn.execute(text('ALTER TABLE test_scenarios ADD COLUMN business_rule_id INTEGER'))
            conn.commit()
        print("Migration: business_rule_id column added successfully!")
    
    # Fix NOT NULL constraint on mapping_id for SQLite
    # SQLite doesn't support ALTER COLUMN, so we need to recreate the table
    print("Checking mapping_id constraint...")
    with engine.connect() as conn:
        # Check if we need to fix the constraint by trying to insert a test row
        try:
            # Create a temporary table with the correct schema
            conn.execute(text('''
                CREATE TABLE IF NOT EXISTS test_scenarios_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    mapping_id INTEGER,
                    endpoint_id INTEGER,
                    name VARCHAR NOT NULL,
                    description TEXT,
                    category VARCHAR,
                    status VARCHAR DEFAULT 'pending',
                    request_body TEXT,
                    expected_response TEXT,
                    json_schema TEXT,
                    business_rule_id INTEGER,
                    FOREIGN KEY(mapping_id) REFERENCES mappings(id),
                    FOREIGN KEY(endpoint_id) REFERENCES api_endpoints(id),
                    FOREIGN KEY(business_rule_id) REFERENCES custom_business_rules(id)
                )
            '''))
            
            # Copy existing data
            conn.execute(text('''
                INSERT INTO test_scenarios_new (id, mapping_id, endpoint_id, name, description, category, status, request_body, expected_response, json_schema, business_rule_id)
                SELECT id, mapping_id, endpoint_id, name, description, category, status, request_body, expected_response, json_schema, 
                       CASE WHEN EXISTS(SELECT 1 FROM pragma_table_info('test_scenarios') WHERE name='business_rule_id') 
                            THEN business_rule_id ELSE NULL END
                FROM test_scenarios
            '''))
            
            # Drop old table and rename new one
            conn.execute(text('DROP TABLE test_scenarios'))
            conn.execute(text('ALTER TABLE test_scenarios_new RENAME TO test_scenarios'))
            conn.commit()
            print("Migration: mapping_id constraint fixed successfully!")
        except Exception as inner_e:
            # If table already has correct schema, this will fail - that's OK
            if "test_scenarios_new" in str(inner_e):
                conn.rollback()
                print("Migration: Schema already correct, skipping...")
            else:
                raise
except Exception as e:
    print(f"Migration warning: {str(e)}")
    import traceback
    traceback.print_exc()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

class MappingResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    id: int
    filename: str
    upload_date: datetime
    status: str
    parsed_count: int

class ParsedFieldResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    id: int
    mapping_id: int
    field_name: str
    data_type: Optional[str]
    required: bool
    example_value: Optional[str]
    gs_rule: Optional[str]
    rule_type: Optional[str]
    confidence: Optional[float]
    routing: Optional[str]
    approved: bool

class TestScenarioResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    id: int
    mapping_id: Optional[int]
    endpoint_id: Optional[int]
    name: str
    description: Optional[str]
    category: str
    status: str
    request_body: Optional[str]
    expected_response: Optional[str]
    json_schema: Optional[str]
    business_rule_id: Optional[int] = None

class TestScenarioUpdate(BaseModel):
    endpoint_id: Optional[int] = None
    request_body: Optional[str] = None
    expected_response: Optional[str] = None
    json_schema: Optional[str] = None

class TestExecutionResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    id: int
    scenario_id: int
    execution_date: datetime
    status: str
    pass_count: int
    fail_count: int
    total_response_time_ms: Optional[int]
    request_body: Optional[str]
    response_body: Optional[str]
    expected_response: Optional[str]

class ValidationResultResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    id: int
    scenario_id: int
    field_name: str
    expected: Optional[str]
    actual: Optional[str]
    status: str
    root_cause: Optional[str]
    validation_type: Optional[str]
    response_time_ms: Optional[int]
    status_code: Optional[int]
    timestamp: datetime

class EnvironmentCreate(BaseModel):
    name: str
    description: Optional[str] = None
    variables: Optional[str] = None

class EnvironmentResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    id: int
    name: str
    description: Optional[str]
    variables: Optional[str]
    is_active: bool
    created_date: datetime

class APIEndpointCreate(BaseModel):
    name: str
    environment_id: int
    base_url: str
    method: str = "GET"
    path: str = "/"
    auth_type: Optional[str] = None
    auth_token: Optional[str] = None
    headers: Optional[str] = None
    default_request_body: Optional[str] = None
    timeout_ms: int = 5000
    expected_status: int = 200
    max_response_time_ms: int = 2000
    token_endpoint_id: Optional[int] = None
    token_response_path: Optional[str] = "access_token"

class APIEndpointResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    id: int
    name: str
    environment_id: int
    base_url: str
    method: str
    path: str
    auth_type: Optional[str]
    auth_token: Optional[str]
    headers: Optional[str]
    default_request_body: Optional[str]
    timeout_ms: int
    expected_status: int
    token_endpoint_id: Optional[int]
    token_response_path: Optional[str]
    max_response_time_ms: int
    created_date: datetime

class CustomBusinessRuleRequest(BaseModel):
    endpoint_id: int
    rule_name: str
    rule_text: str

class CustomBusinessRuleResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    id: int
    endpoint_id: int
    rule_name: str
    rule_text: str
    created_date: datetime
    updated_date: datetime
    is_active: bool

class BusinessRuleValidationResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    id: int
    rule_id: int
    endpoint_id: int
    execution_date: datetime
    record_number: Optional[int]
    field_name: Optional[str]
    expected_value: Optional[str]
    actual_value: Optional[str]
    result: str
    error_message: Optional[str]
    rule_text: Optional[str]

def classify_rule_with_ai(rule_text: str, client, deployment_name: str) -> dict:
    """Classify a single validation rule using OpenAI."""
    import json
    
    prompt = f"""This is a validation rule: '{rule_text}'. Classify as: enum/pattern/conditional/text.
Return JSON: {{"type":"...","confidence":0-100}}"""
    
    try:
        response = client.chat.completions.create(
            model=deployment_name,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=100,
            temperature=0.3
        )
        
        response_text = response.choices[0].message.content.strip()
        
        start_idx = response_text.find('{')
        end_idx = response_text.rfind('}') + 1
        if start_idx != -1 and end_idx > start_idx:
            json_str = response_text[start_idx:end_idx]
            result = json.loads(json_str)
            return {
                "type": result.get("type", "text"),
                "confidence": result.get("confidence", 50)
            }
    except Exception as e:
        print(f"Error classifying rule: {e}")
    
    return {"type": "text", "confidence": 0}

async def test_endpoint_and_get_response(endpoint_url: str, endpoint_obj=None):
    """Test the selected endpoint and return its response for AI analysis."""
    try:
        if not endpoint_url:
            return None
            
        print(f"Testing endpoint: {endpoint_url}")
        
        # Prepare headers
        headers = {}
        
        # Add authentication if endpoint object is provided
        if endpoint_obj:
            # Parse custom headers if provided
            if endpoint_obj.headers:
                try:
                    headers = json.loads(endpoint_obj.headers)
                except json.JSONDecodeError:
                    print(f"Warning: Invalid JSON in endpoint headers")
            
            # Add authentication headers
            if endpoint_obj.auth_type == 'bearer' and endpoint_obj.auth_token:
                headers['Authorization'] = f'Bearer {endpoint_obj.auth_token}'
                print(f"Added Bearer token authentication")
            elif endpoint_obj.auth_type == 'api_key' and endpoint_obj.auth_token:
                headers['X-API-Key'] = endpoint_obj.auth_token
                print(f"Added API Key authentication")
        
        # Make request to the endpoint
        async with httpx.AsyncClient(timeout=10.0, verify=False) as client:
            method = endpoint_obj.method.upper() if endpoint_obj and endpoint_obj.method else 'GET'
            
            if method == 'GET':
                response = await client.get(endpoint_url, headers=headers)
            elif method == 'POST':
                response = await client.post(endpoint_url, headers=headers)
            else:
                response = await client.request(method, endpoint_url, headers=headers)
            
            response_data = {
                "status_code": response.status_code,
                "headers": dict(response.headers),
                "body": response.text[:1000] if response.text else "",  # Limit body size
                "content_type": response.headers.get("content-type", ""),
                "response_size": len(response.content)
            }
            
            print(f"Endpoint response: {response.status_code} - {len(response.content)} bytes")
            return response_data
            
    except Exception as e:
        print(f"Error testing endpoint {endpoint_url}: {str(e)}")
        return {
            "error": str(e),
            "status_code": None,
            "headers": {},
            "body": "",
            "content_type": "",
            "response_size": 0
        }

async def generate_test_scenarios_ai(mapping_id: int, db: Session):
    """Generate comprehensive test scenarios using AI."""
    import json
    
    try:
        mapping = db.query(Mapping).filter(Mapping.id == mapping_id).first()
        if not mapping:
            raise HTTPException(status_code=404, detail="Mapping not found")
        
        parsed_fields = db.query(ParsedField).filter(ParsedField.mapping_id == mapping_id).all()
        
        if not parsed_fields:
            raise HTTPException(status_code=400, detail="No parsed fields found")
        
        azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
        api_key = os.getenv("AZURE_OPENAI_API_KEY")
        api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")
        deployment_name = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "gpt-4")
        
        if not api_key or not azure_endpoint:
            raise HTTPException(status_code=500, detail="Azure OpenAI credentials not configured")
        
        client = AzureOpenAI(
            api_key=api_key,
            api_version=api_version,
            azure_endpoint=azure_endpoint
        )
        
        # Build field information with validation rules and expected responses
        fields_json = []
        has_expected_responses = False
        sample_expected_response = None
        
        for f in parsed_fields:
            field_info = {
                "field_name": f.field_name,
                "data_type": f.data_type,
                "required": f.required,
                "example_value": f.example_value,
                "validation_rule": f.gs_rule if f.gs_rule else None,
                "expected_response": f.expected_response if f.expected_response else None
            }
            fields_json.append(field_info)
            
            # Check if we have any expected responses provided
            if f.expected_response and not has_expected_responses:
                has_expected_responses = True
                sample_expected_response = f.expected_response
        
        # Build prompt with or without expected response guidance
        response_guidance = ""
        if has_expected_responses and sample_expected_response:
            response_guidance = f"""
**IMPORTANT: Use the provided expected_response samples from the field definitions as a guide for the response format.**
Sample expected response format: {sample_expected_response}
"""
        
        # Add business rules if provided
        business_rules_guidance = ""
        business_rules_scenarios = ""
        if mapping.business_rules:
            business_rules_guidance = f"""
**CUSTOM BUSINESS RULES:**
{mapping.business_rules}

**CRITICAL PRIORITY: These business rules are the HIGHEST PRIORITY. Generate test scenarios for EACH business rule FIRST before any other scenarios.**
"""
            # Count business rules to determine how many scenarios to generate for them
            business_rules_scenarios = """
**BUSINESS RULE SCENARIOS (PRIORITY 1-5):**
Generate 5 scenarios specifically for validating the custom business rules provided above. Each scenario should:
- Test a specific business rule
- Include both valid and invalid cases for that rule
- Have clear expected behavior when the rule is violated
- Use realistic data that demonstrates the rule enforcement
"""
        
        # Test selected endpoint and include response analysis
        endpoint_guidance = ""
        if mapping.selected_endpoint:
            print(f"Testing selected endpoint: {mapping.selected_endpoint}")
            # Find the endpoint object to get authentication details
            endpoint_obj = db.query(APIEndpoint).filter(
                (APIEndpoint.base_url + APIEndpoint.path) == mapping.selected_endpoint
            ).first()
            endpoint_response = await test_endpoint_and_get_response(mapping.selected_endpoint, endpoint_obj)
            
            if endpoint_response:
                endpoint_guidance = f"""
**SELECTED ENDPOINT ANALYSIS:**
Endpoint URL: {mapping.selected_endpoint}
Response Status: {endpoint_response.get('status_code', 'Unknown')}
Content-Type: {endpoint_response.get('content_type', 'Unknown')}
Response Size: {endpoint_response.get('response_size', 0)} bytes

**Response Body Sample:**
{endpoint_response.get('body', 'No response body')[:500]}

**IMPORTANT: Use this endpoint information to create realistic test scenarios. Consider the actual response format, status codes, and data structure when generating test cases.**
"""
                if endpoint_response.get('error'):
                    endpoint_guidance += f"""
**ENDPOINT ERROR:**
{endpoint_response['error']}

**NOTE: The endpoint could not be reached. Generate test scenarios based on the expected API behavior and field definitions.**
"""
        
        # Build scenario order based on whether business rules are provided
        if mapping.business_rules:
            scenario_order = f"""
**SCENARIO GENERATION PRIORITY:**

1. **SCHEMA CHECK SCENARIO (1 scenario only):**
   - Create ONE comprehensive schema validation scenario
   - Tests that the response structure matches expected format
   - Validates all required fields are present
   - Category: "positive"

2. **BUSINESS RULE SCENARIOS (15-20 scenarios):**
   - Create dedicated test scenarios for EACH business rule provided
   - For each rule, create both positive (compliant) and negative (violation) test cases
   - Use realistic data from the endpoint response
   - Clearly state which business rule is being tested in the scenario name
   - Category: "positive" for compliant cases, "negative" for violations

3. **FIELD VALIDATION SCENARIOS (10-15 scenarios):**
   - Test each field's data type, format, and constraints
   - Required field validations (missing, null, empty)
   - Data type mismatches
   - Boundary values and length constraints
   - Invalid formats (dates, emails, etc.)
   - Category: mostly "negative" with some "positive" edge cases

4. **EDGE CASE SCENARIOS (5-8 scenarios):**
   - Special characters and encoding
   - Concurrent field updates
   - Optional field combinations
   - Category: mix of "positive" and "negative"

**IMPORTANT:** Generate AT LEAST 30-40 scenarios total. Focus heavily on business rule validation scenarios.
"""
        else:
            scenario_order = """
**SCENARIO GENERATION PRIORITY:**

1. **SCHEMA CHECK SCENARIO (1 scenario only):**
   - Create ONE comprehensive schema validation scenario
   - Tests that the response structure matches expected format
   - Validates all required fields are present
   - Category: "positive"

2. **FIELD VALIDATION SCENARIOS (15-20 scenarios):**
   - Test each field's data type, format, and constraints
   - Required field validations (missing, null, empty)
   - Data type mismatches
   - Boundary values and length constraints
   - Invalid formats (dates, emails, etc.)
   - Category: mostly "negative" with some "positive" edge cases

3. **POSITIVE SCENARIOS (10-12 scenarios):**
   - Happy path with all required fields
   - Happy path with all fields populated
   - Minimal valid request
   - Valid variations of optional fields
   - Valid boundary values
   - Valid format variations

4. **EDGE CASE SCENARIOS (5-8 scenarios):**
   - Special characters and encoding
   - Authentication scenarios
   - Header validation
   - Optional field combinations
   - Category: mix of "positive" and "negative"

**IMPORTANT:** Generate AT LEAST 30-40 scenarios total. Create only ONE schema check scenario.
"""
        
        prompt = f"""You are an API testing expert. Generate comprehensive test scenarios based on the following information:

**FIELD DEFINITIONS:**
{json.dumps(fields_json, indent=2)}

{response_guidance}
{business_rules_guidance}
{endpoint_guidance}

**INSTRUCTIONS:**
1. **Use Realistic Data:** Analyze the endpoint response and field definitions to create realistic test data
2. **Leverage Business Rules:** If business rules are provided, create dedicated scenarios validating each rule (both positive compliance and negative violations)
3. **Balance Positive/Negative:** Generate 40-50% POSITIVE scenarios (category: "positive") and 50-60% NEGATIVE scenarios (category: "negative")
4. **Field-Specific Tests:** Create scenarios specific to each field's data type, validation rules, and expected responses
5. **Complete Request Bodies:** Every scenario must have a COMPLETE request body with ALL fields (use null/omit for testing missing fields)
6. **Realistic Expected Responses:** Base expected responses on the actual endpoint response format shown above

**SCENARIO REQUIREMENTS:**
{scenario_order}

Return ONLY a JSON array with 30-40+ test scenario objects. Each object MUST have:
{{
  "name": "descriptive scenario name",
  "description": "clear explanation of what this test validates (1-2 sentences)",
  "category": "positive" or "negative",
  "request_body": {{complete JSON object with all test fields}},
  "expected_response": {{expected API response JSON}},
  "expected_status": 200 or 400
}}

Example format:
[
  {{
    "name": "Happy Path - All Valid Fields",
    "description": "Validates that the API accepts a request with all required fields populated with valid data and returns a successful response.",
    "category": "positive",
    "request_body": {{"field1": "value1", "field2": 123}},
    "expected_response": {{"success": true, "message": "Created successfully"}},
    "expected_status": 200
  }},
  ...
]

**CRITICAL JSON SYNTAX RULES:**
1. Return ONLY valid JSON - no comments, no markdown, no explanatory text
2. Do NOT use JavaScript code like .repeat(), .map(), etc. - use actual values
3. Do NOT use template literals or string interpolation
4. All string values must be actual strings, not code that generates strings
5. For long repeated strings, write them out fully or use shorter examples
6. Ensure all commas are properly placed (no trailing commas)
7. Ensure all quotes are properly escaped
8. Do NOT include any text before [ or after ]

**EXAMPLE OF INVALID JSON (DO NOT DO THIS):**
{{"name": "A".repeat(100)}}  ❌ WRONG - this is JavaScript code

**EXAMPLE OF VALID JSON (DO THIS):**
{{"name": "AAAAAAAAAA..."}}  ✓ CORRECT - actual string value

Return pure JSON array only, starting with [ and ending with ]."""

        response = client.chat.completions.create(
            model=deployment_name,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=24000,
            temperature=0.3
        )
        
        response_text = response.choices[0].message.content
        print(f"AI Response length: {len(response_text)} characters")
        print(f"AI Response preview: {response_text[:300]}...")
        
        start_idx = response_text.find('[')
        end_idx = response_text.rfind(']') + 1
        
        if start_idx == -1 or end_idx <= start_idx:
            print("ERROR: No JSON array found in AI response")
            raise HTTPException(status_code=500, detail="AI did not return valid JSON array")
        
        json_str = response_text[start_idx:end_idx]
        print(f"Extracted JSON length: {len(json_str)} characters")
        
        # Clean up common JSON issues
        import re
        
        # Remove markdown code blocks
        json_str = json_str.replace('```json', '').replace('```', '')
        
        # Remove JavaScript-style comments (// and /* */)
        # Remove single-line comments
        json_str = re.sub(r'//.*?(?=\n|$)', '', json_str)
        # Remove multi-line comments
        json_str = re.sub(r'/\*.*?\*/', '', json_str, flags=re.DOTALL)
        
        # Fix common AI formatting issues
        # Remove trailing commas before closing brackets
        json_str = re.sub(r',\s*}', '}', json_str)
        json_str = re.sub(r',\s*]', ']', json_str)
        
        # Remove control characters that break JSON parsing
        # Replace unescaped newlines, tabs, and other control characters within strings
        # This regex finds strings and replaces control characters inside them
        def clean_string_content(match):
            string_content = match.group(0)
            # Replace control characters with escaped versions
            string_content = string_content.replace('\n', '\\n')
            string_content = string_content.replace('\r', '\\r')
            string_content = string_content.replace('\t', '\\t')
            # Remove other control characters (ASCII 0-31 except those we just handled)
            string_content = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', string_content)
            return string_content
        
        # Apply cleaning to string values in JSON
        # This pattern matches JSON string values
        json_str = re.sub(r'"[^"\\]*(?:\\.[^"\\]*)*"', clean_string_content, json_str)
        
        try:
            scenarios = json.loads(json_str)
            print(f"Successfully parsed {len(scenarios)} scenarios from AI response")
        except json.JSONDecodeError as e:
            print(f"JSON Parse Error: {e}")
            print(f"Error at position: {e.pos}")
            print(f"Problematic JSON section: ...{json_str[max(0, e.pos-200):min(len(json_str), e.pos+200)]}...")
            
            # Try to salvage partial scenarios
            try:
                # Find the last complete object before the error
                partial_json = json_str[:e.pos]
                
                # Find the last complete '}' that closes an object
                last_complete = partial_json.rfind('}')
                if last_complete != -1:
                    partial_json = partial_json[:last_complete + 1]
                    
                    # Ensure it ends with ]
                    if not partial_json.rstrip().endswith(']'):
                        # Remove trailing comma if present
                        if partial_json.rstrip().endswith(','):
                            partial_json = partial_json.rstrip()[:-1]
                        partial_json = partial_json + ']'
                    
                    scenarios = json.loads(partial_json)
                    print(f"✓ Recovered {len(scenarios)} complete scenarios from partial JSON")
                else:
                    raise Exception("Could not find complete scenarios")
            except Exception as recovery_error:
                print(f"Failed to recover scenarios: {recovery_error}")
                # Save the problematic response for debugging
                with open("failed_ai_response.txt", "w", encoding="utf-8") as f:
                    f.write(response_text)
                print("Saved failed AI response to failed_ai_response.txt")
                raise HTTPException(status_code=500, detail=f"Failed to parse AI response: {str(e)}")
        
        # No longer limiting scenarios - save all generated scenarios
        print(f"AI returned {len(scenarios)} scenarios, saving all of them")
        
        # Use the selected endpoint from the mapping if available
        endpoint_id = None
        if mapping.selected_endpoint:
            # Find the endpoint that matches the selected URL
            endpoint = db.query(APIEndpoint).filter(
                (APIEndpoint.base_url + APIEndpoint.path) == mapping.selected_endpoint
            ).first()
            
            if endpoint:
                endpoint_id = endpoint.id
                print(f"Using selected endpoint: {mapping.selected_endpoint} (ID: {endpoint_id})")
            else:
                print(f"Warning: Selected endpoint not found in database: {mapping.selected_endpoint}")
                # Try to find by partial match
                endpoints = db.query(APIEndpoint).all()
                for ep in endpoints:
                    full_url = ep.base_url + ep.path
                    if full_url == mapping.selected_endpoint:
                        endpoint_id = ep.id
                        print(f"Found endpoint by full URL match: {full_url} (ID: {endpoint_id})")
                        break
        
        if not endpoint_id:
            # Fallback to first available endpoint
            endpoint = db.query(APIEndpoint).first()
            endpoint_id = endpoint.id if endpoint else None
            if endpoint_id:
                print(f"No selected endpoint, using first available: {endpoint.name} (ID: {endpoint_id})")
        
        for scenario_data in scenarios:
            # Convert request_body and expected_response to JSON strings
            request_body_json = json.dumps(scenario_data.get("request_body", {}))
            expected_response_json = json.dumps(scenario_data.get("expected_response", {}))
            
            scenario = TestScenario(
                mapping_id=mapping_id,
                endpoint_id=endpoint_id,
                name=scenario_data.get("name", "Unnamed Scenario"),
                description=scenario_data.get("description", ""),
                category=scenario_data.get("category", "validation"),
                status="pending",
                request_body=request_body_json,
                expected_response=expected_response_json
            )
            db.add(scenario)
        
        db.commit()
        print(f"Generated {len(scenarios)} scenarios with complete request/response JSON")
        return len(scenarios)
        
    except Exception as e:
        print(f"Error generating scenarios: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating scenarios: {str(e)}")

def get_or_fetch_token(endpoint_id: int, db: Session) -> str:
    """Get cached token or fetch new one from token endpoint."""
    import requests
    from datetime import datetime, timedelta
    
    try:
        endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == endpoint_id).first()
        if not endpoint or not endpoint.token_endpoint_id:
            return None
        
        # Check if we have a valid cached token
        now = datetime.utcnow()
        cached_token = db.query(TokenCache).filter(
            TokenCache.endpoint_id == endpoint.token_endpoint_id,
            TokenCache.expires_at > now
        ).order_by(TokenCache.created_at.desc()).first()
        
        if cached_token:
            print(f"Using cached token for endpoint {endpoint_id}")
            return cached_token.token
        
        # Fetch new token from token endpoint
        token_endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == endpoint.token_endpoint_id).first()
        if not token_endpoint:
            print(f"Token endpoint {endpoint.token_endpoint_id} not found")
            return None
        
        print(f"Fetching new token from {token_endpoint.name}")
        
        # Make request to token endpoint
        url = f"{token_endpoint.base_url}{token_endpoint.path}"
        headers = {}
        
        if token_endpoint.headers:
            headers = json.loads(token_endpoint.headers)
        
        body_data = None
        request_body = None
        content_type = headers.get('Content-Type', '').lower()
        
        if token_endpoint.default_request_body:
            body_str = token_endpoint.default_request_body
            if 'application/x-www-form-urlencoded' in content_type:
                body_data = body_str
            else:
                try:
                    request_body = json.loads(body_str)
                except:
                    body_data = body_str
        
        # Make token request
        if body_data:
            response = requests.request(
                method=token_endpoint.method,
                url=url,
                headers=headers,
                data=body_data,
                timeout=token_endpoint.timeout_ms/1000
            )
        elif request_body:
            response = requests.request(
                method=token_endpoint.method,
                url=url,
                headers=headers,
                json=request_body,
                timeout=token_endpoint.timeout_ms/1000
            )
        else:
            response = requests.request(
                method=token_endpoint.method,
                url=url,
                headers=headers,
                timeout=token_endpoint.timeout_ms/1000
            )
        
        if response.status_code != 200:
            print(f"Token request failed: {response.status_code}")
            return None
        
        token_response = response.json()
        
        # Extract token from response using token_response_path
        token_path = endpoint.token_response_path or "access_token"
        token = token_response.get(token_path)
        
        if not token:
            print(f"Token not found in response at path: {token_path}")
            return None
        
        # Calculate expiration time
        expires_in = token_response.get('expires_in', 1800)  # Default 30 minutes
        expires_at = datetime.utcnow() + timedelta(seconds=expires_in - 60)  # Refresh 1 min early
        
        # Cache the token
        new_cache = TokenCache(
            endpoint_id=endpoint.token_endpoint_id,
            token=token,
            expires_at=expires_at
        )
        db.add(new_cache)
        db.commit()
        
        print(f"Token cached, expires in {expires_in} seconds")
        return token
        
    except Exception as e:
        print(f"Error fetching token: {e}")
        return None

ROOT_CAUSE_CATEGORIES = {
    'API_BUG': 'API Bug',
    'DATA_TRANSFORMATION': 'Data Transformation Error',
    'VENDOR_DATA': 'Vendor Data Issue',
    'STALE_CACHE': 'Stale Cache',
    'BUSINESS_RULE': 'Business Rule Violation',
    'MISSING_SYNC': 'Missing Sync',
    'CONFIGURATION': 'Configuration Error'
}

def get_root_cause_from_ai(field_name: str, expected: str, actual: str, validation_result_id: int = None, db: Session = None) -> str:
    """Get root cause explanation from Azure OpenAI for validation failure.
    This is a simplified version for backward compatibility. Use analyze_failure_with_ai for full analysis."""
    import json
    
    try:
        azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
        api_key = os.getenv("AZURE_OPENAI_API_KEY")
        api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")
        deployment_name = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "gpt-4")
        
        if not api_key or not azure_endpoint:
            return "Azure OpenAI not configured"
        
        client = AzureOpenAI(
            api_key=api_key,
            api_version=api_version,
            azure_endpoint=azure_endpoint
        )
        
        prompt = f"""Field {field_name} failed. Expected: {expected}. Actual: {actual}.
Explain why in 1 sentence."""
        
        response = client.chat.completions.create(
            model=deployment_name,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=100,
            temperature=0.3
        )
        
        return response.choices[0].message.content.strip()
        
    except Exception as e:
        print(f"Error getting root cause: {e}")
        return f"Failed to analyze: {str(e)}"

def analyze_failure_with_ai(validation_result_id: int, db: Session):
    """Comprehensive AI analysis of a validation failure with context gathering and caching."""
    import json
    import time
    
    try:
        # Fetch the failed validation result
        validation_result = db.query(ValidationResult).filter(ValidationResult.id == validation_result_id).first()
        
        if not validation_result:
            return {"error": "Validation result not found"}
        
        if validation_result.status != 'fail':
            return {"error": "This validation passed, no analysis needed"}
        
        # Check cache first
        cache_key = f"{validation_result.field_name}:{validation_result.expected}:{validation_result.actual}"
        cached_analysis = db.query(RootCauseCache).filter(RootCauseCache.cache_key == cache_key).first()
        
        if cached_analysis:
            print(f"Using cached analysis for {validation_result.field_name}")
            return {
                'root_cause_category': cached_analysis.category,
                'explanation': cached_analysis.explanation,
                'suggested_action': cached_analysis.suggested_action,
                'confidence': cached_analysis.confidence,
                'from_cache': True
            }
        
        # Get scenario and mapping context
        scenario = db.query(TestScenario).filter(TestScenario.id == validation_result.scenario_id).first()
        if not scenario:
            return {"error": "Scenario not found"}
        
        mapping_id = scenario.mapping_id
        
        # Get parsed field metadata
        parsed_field = db.query(ParsedField).filter(
            ParsedField.mapping_id == mapping_id,
            ParsedField.field_name == validation_result.field_name
        ).first()
        
        # Get related fields from the same execution
        related_validations = db.query(ValidationResult).filter(
            ValidationResult.execution_id == validation_result.execution_id,
            ValidationResult.field_name != validation_result.field_name
        ).limit(10).all()
        
        related_fields = {}
        for v in related_validations:
            try:
                # Try to parse as JSON, otherwise use as string
                actual_val = json.loads(v.actual) if v.actual and v.actual.startswith('{') else v.actual
                related_fields[v.field_name] = actual_val
            except:
                related_fields[v.field_name] = v.actual
        
        # Get historical pattern
        past_results = db.query(ValidationResult).filter(
            ValidationResult.field_name == validation_result.field_name
        ).order_by(ValidationResult.timestamp.desc()).limit(10).all()
        
        pass_count = sum(1 for r in past_results if r.status == 'pass')
        fail_count = len(past_results) - pass_count
        pass_rate = int((pass_count / len(past_results)) * 100) if past_results else 0
        
        common_failures = {}
        for r in past_results:
            if r.status == 'fail' and r.actual:
                val = r.actual[:50]  # Truncate long values
                common_failures[val] = common_failures.get(val, 0) + 1
        
        history = {
            'pass_rate': pass_rate,
            'fail_count': fail_count,
            'total_runs': len(past_results),
            'common_failures': list(common_failures.keys())[:3]
        }
        
        # Build comprehensive AI prompt
        field_metadata = ""
        if parsed_field:
            field_metadata = f"""
FIELD METADATA:
- Data Type: {parsed_field.data_type}
- Required: {'Yes' if parsed_field.required else 'No'}
- Validation Rule: {parsed_field.gs_rule if parsed_field.gs_rule else 'None'}
- Example Value: {parsed_field.example_value if parsed_field.example_value else 'None'}
"""
        
        related_fields_str = ""
        if related_fields:
            related_fields_str = f"""
RELATED FIELD VALUES (from same API response):
{json.dumps(related_fields, indent=2)[:500]}
"""
        
        prompt = f"""You are analyzing an API validation failure. Provide a concise root cause analysis.

FAILURE DETAILS:
- Field: {validation_result.field_name}
- Validation Type: {validation_result.validation_type}
- Expected: {validation_result.expected}
- Actual: {validation_result.actual}
{field_metadata}{related_fields_str}
HISTORICAL PATTERN (last {history['total_runs']} runs):
- Pass rate: {history['pass_rate']}%
- This field fails {history['fail_count']} of last {history['total_runs']} runs
- Common failure values: {', '.join(history['common_failures']) if history['common_failures'] else 'N/A'}

ANALYSIS TASK:
Provide a 2-3 sentence root cause analysis that:
1. Explains WHY this validation failed
2. Identifies the MOST LIKELY root cause from these categories:
   - API_BUG: API is returning incorrect data
   - DATA_TRANSFORMATION: Mapping/conversion issue between systems
   - VENDOR_DATA: Test data was not created correctly
   - STALE_CACHE: API is returning cached/outdated data
   - BUSINESS_RULE: Field violates a conditional rule
   - MISSING_SYNC: Related fields are out of sync
   - CONFIGURATION: API endpoint or parameters incorrect

3. Suggests WHAT TO CHECK NEXT (specific, actionable)

Return ONLY valid JSON (no markdown, no code blocks):
{{
  "root_cause_category": "one of the 7 categories above",
  "explanation": "2-3 sentence plain-English explanation",
  "suggested_action": "Specific next step to investigate",
  "confidence": 0-100
}}

Focus on being SPECIFIC and ACTIONABLE, not generic."""
        
        # Call Azure OpenAI
        azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
        api_key = os.getenv("AZURE_OPENAI_API_KEY")
        api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")
        deployment_name = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "gpt-4")
        
        if not api_key or not azure_endpoint:
            return {"error": "Azure OpenAI not configured"}
        
        client = AzureOpenAI(
            api_key=api_key,
            api_version=api_version,
            azure_endpoint=azure_endpoint
        )
        
        response = client.chat.completions.create(
            model=deployment_name,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=500,
            temperature=0.3
        )
        
        # Parse AI response
        response_text = response.choices[0].message.content.strip()
        
        # Clean up response (remove markdown code blocks if present)
        if response_text.startswith('```'):
            response_text = response_text.split('```')[1]
            if response_text.startswith('json'):
                response_text = response_text[4:]
            response_text = response_text.strip()
        
        ai_analysis = json.loads(response_text)
        
        # Cache the analysis
        cache_entry = RootCauseCache(
            cache_key=cache_key,
            category=ai_analysis.get('root_cause_category', 'UNKNOWN'),
            explanation=ai_analysis.get('explanation', ''),
            suggested_action=ai_analysis.get('suggested_action', ''),
            confidence=ai_analysis.get('confidence', 50)
        )
        db.add(cache_entry)
        
        # Update validation result with AI analysis
        validation_result.root_cause = ai_analysis.get('explanation', '')
        validation_result.root_cause_category = ai_analysis.get('root_cause_category', 'UNKNOWN')
        validation_result.suggested_action = ai_analysis.get('suggested_action', '')
        validation_result.ai_confidence = ai_analysis.get('confidence', 50)
        
        db.commit()
        
        ai_analysis['from_cache'] = False
        return ai_analysis
        
    except Exception as e:
        print(f"Error in AI analysis: {e}")
        import traceback
        traceback.print_exc()
        return {"error": f"Analysis failed: {str(e)}"}

def validate_json_schema(data: dict, schema: dict) -> tuple:
    """Validate JSON data against schema. Returns (is_valid, errors)."""
    try:
        from jsonschema import validate, ValidationError
        validate(instance=data, schema=schema)
        return True, []
    except ValidationError as e:
        return False, [str(e.message)]
    except Exception as e:
        return False, [f"Schema validation error: {str(e)}"]

def get_value_by_path(payload: dict, path: str):
    """Get nested value from dict using dot-separated path."""
    if not path:
        return None
    current = payload
    for part in path.split('.'):
        if not isinstance(current, dict) or part not in current:
            return None
        current = current[part]
    return current

def normalize_text_content(value: str) -> str:
    """Normalize text content by stripping HTML and normalizing whitespace."""
    if value is None:
        return ""
    import re
    text = str(value)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def normalize_cell_value(value):
    if value is None:
        return ""
    return str(value).strip()

def compare_excel_row_to_record(row_data: dict, record: dict, text_content_field: str = "textContent"):
    """Compare excel row to a single API record."""
    field_results = []
    for field_name, expected_value in row_data.items():
        actual_value = record.get(field_name)
        if field_name == text_content_field:
            expected_norm = normalize_text_content(expected_value)
            actual_norm = normalize_text_content(actual_value)
            status = "pass" if expected_norm == actual_norm else "fail"
            message = "" if status == "pass" else "Normalized text mismatch"
            field_results.append({
                "field": field_name,
                "expected": expected_norm,
                "actual": actual_norm,
                "status": status,
                "message": message
            })
        else:
            expected_norm = normalize_cell_value(expected_value)
            actual_norm = normalize_cell_value(actual_value)
            status = "pass" if expected_norm == actual_norm else "fail"
            message = "" if status == "pass" else "Value mismatch"
            field_results.append({
                "field": field_name,
                "expected": expected_norm,
                "actual": actual_norm,
                "status": status,
                "message": message
            })
    return field_results

def execute_validation(scenario_id: int, db: Session, request_body: dict = None):
    """Execute validation with 5 types: SCHEMA, TYPE, STATUS_CODE, RESPONSE_TIME, and JSON_SCHEMA checks."""
    import json
    import requests
    import time
    
    try:
        scenario = db.query(TestScenario).filter(TestScenario.id == scenario_id).first()
        if not scenario:
            raise HTTPException(status_code=404, detail="Scenario not found")
        
        scenario.status = "running"
        db.commit()
        
        # Create TestExecution record
        execution = TestExecution(
            scenario_id=scenario_id,
            status="running",
            request_body=json.dumps(request_body) if request_body else None
        )
        db.add(execution)
        db.commit()
        db.refresh(execution)
        
        parsed_fields = db.query(ParsedField).filter(
            ParsedField.mapping_id == scenario.mapping_id
        ).all()
        
        # Use provided request_body or scenario's request_body or create mock data
        if not request_body:
            if scenario.request_body:
                try:
                    request_body = json.loads(scenario.request_body)
                except:
                    request_body = {field.field_name: field.example_value for field in parsed_fields if field.example_value}
            else:
                request_body = {field.field_name: field.example_value for field in parsed_fields if field.example_value}
        
        # Make actual API call if endpoint is configured
        api_response = None
        response_time_ms = None
        status_code = None
        
        if scenario.endpoint_id:
            endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == scenario.endpoint_id).first()
            if endpoint:
                try:
                    url = f"{endpoint.base_url}{endpoint.path}"
                    headers = {}
                    
                    if endpoint.headers:
                        headers = json.loads(endpoint.headers)
                    
                    # Check if endpoint needs token from another endpoint
                    if endpoint.token_endpoint_id:
                        token = get_or_fetch_token(endpoint.id, db)
                        if token:
                            headers["Authorization"] = f"Bearer {token}"
                            print(f"Using auto-fetched token for {endpoint.name}")
                        else:
                            print(f"Failed to fetch token for {endpoint.name}")
                    elif endpoint.auth_type == "bearer" and endpoint.auth_token:
                        headers["Authorization"] = f"Bearer {endpoint.auth_token}"
                    elif endpoint.auth_type == "api_key" and endpoint.auth_token:
                        headers["X-API-Key"] = endpoint.auth_token
                    
                    start_time = time.time()
                    
                    print(f"Making API call to: {url}")
                    print(f"Method: {endpoint.method}")
                    print(f"Headers: {headers}")
                    
                    if endpoint.method.upper() == "GET":
                        response = requests.get(url, headers=headers, timeout=endpoint.timeout_ms/1000)
                    elif endpoint.method.upper() == "POST":
                        response = requests.post(url, json=request_body, headers=headers, timeout=endpoint.timeout_ms/1000)
                    elif endpoint.method.upper() == "PUT":
                        response = requests.put(url, json=request_body, headers=headers, timeout=endpoint.timeout_ms/1000)
                    elif endpoint.method.upper() == "DELETE":
                        response = requests.delete(url, headers=headers, timeout=endpoint.timeout_ms/1000)
                    else:
                        response = requests.request(endpoint.method, url, json=request_body, headers=headers, timeout=endpoint.timeout_ms/1000)
                    
                    response_time_ms = int((time.time() - start_time) * 1000)
                    status_code = response.status_code
                    
                    print(f"Response Status: {status_code}")
                    print(f"Response Headers: {response.headers}")
                    print(f"Response Text (first 500 chars): {response.text[:500]}")
                    
                    try:
                        api_response = response.json()
                        print(f"Parsed JSON response type: {type(api_response)}")
                        
                        # Store original response for metadata fields
                        original_response = api_response
                        
                        # If response is an array, use the first item for validation
                        if isinstance(api_response, list):
                            if len(api_response) > 0:
                                print(f"Response is array with {len(api_response)} items, using first item")
                                api_response = api_response[0]
                            else:
                                print("Response is empty array")
                                api_response = {}
                        # If response is object with dataset array, extract first item from dataset
                        elif isinstance(api_response, dict):
                            # Try common dataset keys
                            for dataset_key in ['termsAndConditions', 'entityData', 'data', 'results', 'items', 'records']:
                                if dataset_key in api_response and isinstance(api_response[dataset_key], list):
                                    if len(api_response[dataset_key]) > 0:
                                        print(f"Found dataset '{dataset_key}' with {len(api_response[dataset_key])} items")
                                        # Keep both root and first item for field lookup
                                        first_item = api_response[dataset_key][0]
                                        # Merge root fields with first item fields for lookup
                                        api_response = {**api_response, **first_item}
                                        print(f"Merged root fields with first item from {dataset_key}")
                                        break
                        
                        print(f"Final api_response for validation: {api_response}")
                    except Exception as json_error:
                        print(f"Failed to parse JSON: {json_error}")
                        api_response = {"response": response.text}
                    
                except Exception as e:
                    print(f"API call failed: {e}")
                    api_response = None
        
        pass_count = 0
        fail_count = 0
        results = []
        
        for field in parsed_fields:
            field_name = field.field_name
            expected_type = field.data_type
            is_required = field.required
            
            # Get the actual value from API response if available
            response_field_value = None
            if api_response and isinstance(api_response, dict):
                response_field_value = api_response.get(field_name)
            
            # VALIDATION 1: SCHEMA - Check required fields are present in API response
            if is_required and not response_field_value:
                root_cause = get_root_cause_from_ai(
                    field_name,
                    "Field must be present",
                    f"Field '{field_name}' not found in API response"
                )
                
                result = ValidationResult(
                    scenario_id=scenario_id,
                    execution_id=execution.id,
                    field_name=field_name,
                    expected="field present",
                    actual="missing",
                    status="fail",
                    root_cause=root_cause,
                    validation_type="schema"
                )
                db.add(result)
                fail_count += 1
                results.append({
                    "field_name": field_name,
                    "expected": "field present",
                    "actual": "missing",
                    "actual_value": response_field_value,
                    "status": "fail",
                    "root_cause": root_cause
                })
                continue
            
            # Add success case for required fields that are present
            if is_required and response_field_value is not None:
                result = ValidationResult(
                    scenario_id=scenario_id,
                    execution_id=execution.id,
                    field_name=field_name,
                    expected="field present",
                    actual="present",
                    status="pass",
                    validation_type="schema"
                )
                db.add(result)
                pass_count += 1
                results.append({
                    "field_name": field_name,
                    "expected": "field present",
                    "actual": "present",
                    "actual_value": response_field_value,
                    "status": "pass",
                    "root_cause": None
                })
            
            # If field not in request and not required, skip
            if field_name not in request_body:
                continue
            
            actual_value = request_body[field_name]
            
            # VALIDATION 2: TYPE - Check data types match Column B
            type_valid = True
            type_error = None
            
            if expected_type == "string" and not isinstance(actual_value, str):
                type_valid = False
                type_error = f"Expected string, got {type(actual_value).__name__}"
            elif expected_type == "number" and not isinstance(actual_value, (int, float)):
                type_valid = False
                type_error = f"Expected number, got {type(actual_value).__name__}"
            elif expected_type == "boolean" and not isinstance(actual_value, bool):
                type_valid = False
                type_error = f"Expected boolean, got {type(actual_value).__name__}"
            
            if not type_valid:
                root_cause = get_root_cause_from_ai(
                    field_name,
                    f"Type: {expected_type}",
                    type_error
                )
                
                result = ValidationResult(
                    scenario_id=scenario_id,
                    execution_id=execution.id,
                    field_name=field_name,
                    expected=f"Type: {expected_type}",
                    actual=type_error,
                    status="fail",
                    root_cause=root_cause,
                    validation_type="TYPE"
                )
                db.add(result)
                fail_count += 1
                results.append({
                    "field_name": field_name,
                    "expected": f"Type: {expected_type}",
                    "actual": type_error,
                    "actual_value": response_field_value,
                    "status": "fail",
                    "root_cause": root_cause
                })
            else:
                result = ValidationResult(
                    scenario_id=scenario_id,
                    execution_id=execution.id,
                    field_name=field_name,
                    expected=f"Type: {expected_type}, Required: {is_required}",
                    actual=f"Valid {expected_type}",
                    status="pass",
                    root_cause=None,
                    validation_type="TYPE"
                )
                db.add(result)
                pass_count += 1
                results.append({
                    "field_name": field_name,
                    "expected": f"Type: {expected_type}, Required: {is_required}",
                    "actual": f"Valid {expected_type}",
                    "actual_value": response_field_value,
                    "status": "pass",
                    "root_cause": None
                })
        
        # VALIDATION 3: STATUS CODE - Check if API returned expected status
        if scenario.endpoint_id and status_code is not None:
            endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == scenario.endpoint_id).first()
            if endpoint:
                if status_code != endpoint.expected_status:
                    root_cause = get_root_cause_from_ai(
                        "HTTP Status Code",
                        f"Status {endpoint.expected_status}",
                        f"Status {status_code}"
                    )
                    
                    result = ValidationResult(
                        scenario_id=scenario_id,
                        execution_id=execution.id,
                        field_name="HTTP_STATUS",
                        expected=f"Status Code: {endpoint.expected_status}",
                        actual=f"Status Code: {status_code}",
                        status="fail",
                        root_cause=root_cause,
                        validation_type="STATUS_CODE",
                        status_code=status_code
                    )
                    db.add(result)
                    fail_count += 1
                    results.append({
                        "field_name": "HTTP_STATUS",
                        "expected": f"Status Code: {endpoint.expected_status}",
                        "actual": f"Status Code: {status_code}",
                        "status": "fail",
                        "root_cause": root_cause,
                        "validation_type": "STATUS_CODE"
                    })
                else:
                    result = ValidationResult(
                        scenario_id=scenario_id,
                        execution_id=execution.id,
                        field_name="HTTP_STATUS",
                        expected=f"Status Code: {endpoint.expected_status}",
                        actual=f"Status Code: {status_code}",
                        status="pass",
                        validation_type="STATUS_CODE",
                        status_code=status_code
                    )
                    db.add(result)
                    pass_count += 1
                    results.append({
                        "field_name": "HTTP_STATUS",
                        "expected": f"Status Code: {endpoint.expected_status}",
                        "actual": f"Status Code: {status_code}",
                        "status": "pass",
                        "validation_type": "STATUS_CODE"
                    })
        
        # VALIDATION 4: RESPONSE TIME - Check if API responded within SLA
        if scenario.endpoint_id and response_time_ms is not None:
            endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == scenario.endpoint_id).first()
            if endpoint:
                if response_time_ms > endpoint.max_response_time_ms:
                    root_cause = get_root_cause_from_ai(
                        "Response Time",
                        f"< {endpoint.max_response_time_ms}ms",
                        f"{response_time_ms}ms (exceeded by {response_time_ms - endpoint.max_response_time_ms}ms)"
                    )
                    
                    result = ValidationResult(
                        scenario_id=scenario_id,
                        execution_id=execution.id,
                        field_name="RESPONSE_TIME",
                        expected=f"< {endpoint.max_response_time_ms}ms",
                        actual=f"{response_time_ms}ms",
                        status="fail",
                        root_cause=root_cause,
                        validation_type="RESPONSE_TIME",
                        response_time_ms=response_time_ms
                    )
                    db.add(result)
                    fail_count += 1
                    results.append({
                        "field_name": "RESPONSE_TIME",
                        "expected": f"< {endpoint.max_response_time_ms}ms",
                        "actual": f"{response_time_ms}ms",
                        "status": "fail",
                        "root_cause": root_cause,
                        "validation_type": "RESPONSE_TIME"
                    })
                else:
                    result = ValidationResult(
                        scenario_id=scenario_id,
                        execution_id=execution.id,
                        field_name="RESPONSE_TIME",
                        expected=f"< {endpoint.max_response_time_ms}ms",
                        actual=f"{response_time_ms}ms",
                        status="pass",
                        validation_type="RESPONSE_TIME",
                        response_time_ms=response_time_ms
                    )
                    db.add(result)
                    pass_count += 1
                    results.append({
                        "field_name": "RESPONSE_TIME",
                        "expected": f"< {endpoint.max_response_time_ms}ms",
                        "actual": f"{response_time_ms}ms",
                        "status": "pass",
                        "validation_type": "RESPONSE_TIME"
                    })
        
        # VALIDATION 5: JSON SCHEMA - Validate response against JSON schema
        print(f"JSON Schema Check - scenario.json_schema exists: {bool(scenario.json_schema)}")
        print(f"JSON Schema Check - api_response exists: {bool(api_response)}")
        if scenario.json_schema:
            print(f"JSON Schema length: {len(scenario.json_schema)}")
        if api_response:
            print(f"API Response type: {type(api_response)}")
            print(f"API Response preview: {str(api_response)[:200]}...")
        
        if scenario.json_schema and api_response:
            try:
                schema = json.loads(scenario.json_schema)
                is_valid, errors = validate_json_schema(api_response, schema)
                
                # Store actual response data for display
                actual_response_data = json.dumps(api_response, indent=2, ensure_ascii=False)
                
                if not is_valid:
                    root_cause = get_root_cause_from_ai(
                        "JSON Schema",
                        "Valid JSON schema",
                        f"Schema validation failed: {', '.join(errors)}"
                    )
                    
                    result = ValidationResult(
                        scenario_id=scenario_id,
                        execution_id=execution.id,
                        field_name="JSON_SCHEMA",
                        expected="Valid JSON schema",
                        actual=actual_response_data,
                        status="fail",
                        root_cause=root_cause,
                        validation_type="JSON_SCHEMA"
                    )
                    db.add(result)
                    fail_count += 1
                    results.append({
                        "field_name": "JSON_SCHEMA",
                        "expected": "Valid JSON schema",
                        "actual": actual_response_data,
                        "status": "fail",
                        "root_cause": root_cause,
                        "validation_type": "JSON_SCHEMA"
                    })
                else:
                    result = ValidationResult(
                        scenario_id=scenario_id,
                        execution_id=execution.id,
                        field_name="JSON_SCHEMA",
                        expected="Valid JSON schema",
                        actual=actual_response_data,
                        status="pass",
                        validation_type="JSON_SCHEMA"
                    )
                    db.add(result)
                    pass_count += 1
                    results.append({
                        "field_name": "JSON_SCHEMA",
                        "expected": "Valid JSON schema",
                        "actual": actual_response_data,
                        "status": "pass",
                        "validation_type": "JSON_SCHEMA"
                    })
            except Exception as e:
                print(f"JSON schema validation error: {e}")
        else:
            # Create a result to show why JSON schema validation didn't run
            reason = []
            if not scenario.json_schema:
                reason.append("No JSON schema defined for scenario")
            if not api_response:
                reason.append("No API response available")
            
            result = ValidationResult(
                scenario_id=scenario_id,
                execution_id=execution.id,
                field_name="JSON_SCHEMA",
                expected="Valid JSON schema",
                actual=f"Validation skipped: {', '.join(reason)}",
                status="info",
                validation_type="JSON_SCHEMA"
            )
            db.add(result)
            results.append({
                "field_name": "JSON_SCHEMA",
                "expected": "Valid JSON schema",
                "actual": f"Validation skipped: {', '.join(reason)}",
                "status": "info",
                "validation_type": "JSON_SCHEMA"
            })
        
        # VALIDATION 6: RESPONSE_MATCH - Compare expected response vs actual response
        if scenario.expected_response and api_response is not None:
            try:
                expected_resp = json.loads(scenario.expected_response)
                
                # Deep comparison of expected vs actual response
                def is_timestamp(value):
                    """Check if a string value is a timestamp"""
                    if not isinstance(value, str):
                        return False
                    
                    # Common timestamp patterns
                    timestamp_patterns = [
                        r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}',  # ISO 8601
                        r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}',  # SQL datetime
                        r'\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}',  # US datetime
                    ]
                    
                    import re
                    for pattern in timestamp_patterns:
                        if re.search(pattern, value):
                            return True
                    return False
                
                def parse_timestamp(value):
                    """Parse timestamp string to datetime object"""
                    from dateutil import parser
                    try:
                        return parser.parse(value)
                    except:
                        return None
                
                def compare_timestamps(expected_ts, actual_ts, tolerance_seconds=300):
                    """Compare two timestamps with tolerance (default 5 minutes)"""
                    exp_dt = parse_timestamp(expected_ts)
                    act_dt = parse_timestamp(actual_ts)
                    
                    if exp_dt is None or act_dt is None:
                        return False, "Could not parse timestamps"
                    
                    # Calculate time difference
                    diff = abs((exp_dt - act_dt).total_seconds())
                    
                    if diff <= tolerance_seconds:
                        return True, f"Timestamps within tolerance ({diff:.1f}s difference)"
                    else:
                        return False, f"Timestamps differ by {diff:.1f}s (tolerance: {tolerance_seconds}s)"
                
                def compare_responses(expected, actual, path=""):
                    """Recursively compare expected and actual responses"""
                    differences = []
                    
                    if type(expected) != type(actual):
                        differences.append(f"{path}: Type mismatch - expected {type(expected).__name__}, got {type(actual).__name__}")
                        return differences
                    
                    if isinstance(expected, dict):
                        # Check for missing keys in actual
                        for key in expected:
                            if key not in actual:
                                differences.append(f"{path}.{key}: Missing in actual response")
                            else:
                                differences.extend(compare_responses(expected[key], actual[key], f"{path}.{key}"))
                    elif isinstance(expected, list):
                        if len(expected) != len(actual):
                            differences.append(f"{path}: Array length mismatch - expected {len(expected)}, got {len(actual)}")
                        else:
                            for i, (exp_item, act_item) in enumerate(zip(expected, actual)):
                                differences.extend(compare_responses(exp_item, act_item, f"{path}[{i}]"))
                    else:
                        # Compare primitive values
                        if expected != actual:
                            # Check if both values are timestamps
                            if is_timestamp(str(expected)) and is_timestamp(str(actual)):
                                is_match, message = compare_timestamps(str(expected), str(actual))
                                if not is_match:
                                    differences.append(f"{path}: Timestamp mismatch - {message}")
                                # If timestamps match within tolerance, don't add to differences
                            else:
                                differences.append(f"{path}: Value mismatch - expected '{expected}', got '{actual}'")
                    
                    return differences
                
                differences = compare_responses(expected_resp, api_response, "response")
                
                if differences:
                    diff_summary = "; ".join(differences[:5])  # Show first 5 differences
                    if len(differences) > 5:
                        diff_summary += f" ... and {len(differences) - 5} more differences"
                    
                    root_cause = get_root_cause_from_ai(
                        "Response Match",
                        json.dumps(expected_resp, indent=2),
                        json.dumps(api_response, indent=2)
                    )
                    
                    result = ValidationResult(
                        scenario_id=scenario_id,
                        execution_id=execution.id,
                        field_name="RESPONSE_MATCH",
                        expected=json.dumps(expected_resp),
                        actual=json.dumps(api_response),
                        status="fail",
                        root_cause=root_cause or diff_summary,
                        validation_type="RESPONSE_MATCH"
                    )
                    db.add(result)
                    fail_count += 1
                    results.append({
                        "field_name": "RESPONSE_MATCH",
                        "expected": json.dumps(expected_resp),
                        "actual": json.dumps(api_response),
                        "status": "fail",
                        "root_cause": root_cause or diff_summary,
                        "validation_type": "RESPONSE_MATCH"
                    })
                else:
                    result = ValidationResult(
                        scenario_id=scenario_id,
                        execution_id=execution.id,
                        field_name="RESPONSE_MATCH",
                        expected=json.dumps(expected_resp),
                        actual=json.dumps(api_response),
                        status="pass",
                        validation_type="RESPONSE_MATCH"
                    )
                    db.add(result)
                    pass_count += 1
                    results.append({
                        "field_name": "RESPONSE_MATCH",
                        "expected": json.dumps(expected_resp),
                        "actual": json.dumps(api_response),
                        "status": "pass",
                        "validation_type": "RESPONSE_MATCH"
                    })
            except Exception as e:
                print(f"Response match validation error: {e}")
        
        # Update execution record
        execution.status = "completed"
        execution.pass_count = pass_count
        execution.fail_count = fail_count
        execution.total_response_time_ms = response_time_ms
        execution.response_body = json.dumps(api_response) if api_response else None
        execution.expected_response = scenario.expected_response
        
        scenario.status = "completed"
        db.commit()
        
        return {
            "execution_id": execution.id,
            "pass_count": pass_count,
            "fail_count": fail_count,
            "results": results,
            "response_time_ms": response_time_ms,
            "status_code": status_code,
            "actual_response": api_response if api_response else {},
            "expected_response": json.loads(scenario.expected_response) if scenario.expected_response else {}
        }
        
    except Exception as e:
        if 'execution' in locals():
            execution.status = "failed"
        if 'scenario' in locals():
            scenario.status = "failed"
        db.commit()
        print(f"Error executing validation: {e}")
        raise HTTPException(status_code=500, detail=f"Error executing validation: {str(e)}")

@app.get("/")
def read_root():
    return {"message": "GS API Test Platform"}

@app.post("/api/mappings/upload")
async def upload_mapping(file: UploadFile = File(...), business_rules: str = Form(None), selected_endpoint: str = Form(None)):
    db = next(get_db())
    try:
        print(f"Received file upload: {file.filename}")
        print(f"File content type: {file.content_type}")
        print(f"UPLOAD_DIR exists: {UPLOAD_DIR.exists()}")
        print(f"UPLOAD_DIR path: {UPLOAD_DIR}")
        
        if not file.filename:
            raise HTTPException(status_code=400, detail="No file provided")
            
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files allowed")
        
        # Ensure uploads directory exists
        UPLOAD_DIR.mkdir(exist_ok=True)
        
        file_path = UPLOAD_DIR / file.filename
        print(f"Saving file to: {file_path}")
        
        try:
            content = await file.read()
            print(f"File read successfully: {len(content)} bytes")
            
            if len(content) == 0:
                raise HTTPException(status_code=400, detail="File is empty")
            
            if len(content) > 50 * 1024 * 1024:  # 50MB limit
                raise HTTPException(status_code=400, detail="File too large (max 50MB)")
            
            with open(file_path, "wb") as f:
                f.write(content)
            
            print(f"File saved successfully")
        except Exception as e:
            print(f"Error saving file: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error saving file: {str(e)}")
        
        # Store business rules if provided
        if business_rules:
            print(f"Business rules provided: {business_rules[:100]}...")
        
        # Store selected endpoint if provided
        if selected_endpoint:
            print(f"Selected endpoint: {selected_endpoint}")
        
        mapping = Mapping(filename=file.filename, status="parsing", business_rules=business_rules, selected_endpoint=selected_endpoint)
        db.add(mapping)
        db.commit()
        db.refresh(mapping)
        
        print(f"Mapping created with ID: {mapping.id}")
        
        azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
        api_key = os.getenv("AZURE_OPENAI_API_KEY")
        api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")
        deployment_name = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "gpt-4")
        
        client = None
        use_ai = False
        
        if api_key and azure_endpoint:
            try:
                client = AzureOpenAI(
                    api_key=api_key,
                    api_version=api_version,
                    azure_endpoint=azure_endpoint
                )
                use_ai = True
                print("Azure OpenAI client initialized successfully")
            except Exception as e:
                print(f"Warning: Azure OpenAI initialization failed: {e}")
                print("Continuing without AI classification...")
        else:
            print("Azure OpenAI credentials not configured. Parsing without AI classification...")
        
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        fields_parsed = 0
        total_confidence = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not row or not row[0]:
                continue
            
            # Safely extract values with bounds checking
            field_name = str(row[0]) if len(row) > 0 and row[0] else ""
            data_type = str(row[1]) if len(row) > 1 and row[1] else "string"
            required = str(row[2]).upper() == "Y" if len(row) > 2 and row[2] else False
            example_value = str(row[3]) if len(row) > 3 and row[3] else ""
            rule_text = str(row[5]) if len(row) > 5 and row[5] else ""
            expected_response = str(row[6]) if len(row) > 6 and row[6] else ""
            
            rule_classification = {"type": "text", "confidence": 0}
            if rule_text and use_ai and client:
                print(f"Classifying rule for {field_name}: {rule_text[:50]}...")
                rule_classification = classify_rule_with_ai(rule_text, client, deployment_name)
            elif rule_text:
                print(f"Skipping AI classification for {field_name} (AI not configured)")
                rule_classification = {"type": "text", "confidence": 50}
            
            parsed_field = ParsedField(
                mapping_id=mapping.id,
                field_name=field_name,
                data_type=data_type,
                required=required,
                example_value=example_value,
                gs_rule=rule_text,
                expected_response=expected_response,
                rule_type=rule_classification["type"],
                confidence=rule_classification["confidence"] / 100.0,
                routing=None,
                approved=False
            )
            db.add(parsed_field)
            fields_parsed += 1
            total_confidence += rule_classification["confidence"]
        
        avg_confidence = int(total_confidence / fields_parsed) if fields_parsed > 0 else 0
        
        mapping.parsed_count = fields_parsed
        mapping.status = "parsed"
        db.commit()
        
        print(f"Parsed {fields_parsed} fields with avg confidence {avg_confidence}%")
        
        return {
            "mapping_id": mapping.id,
            "fields_parsed": fields_parsed,
            "avg_confidence": avg_confidence
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error uploading file: {e}")
        import traceback
        traceback.print_exc()
        if 'mapping' in locals():
            mapping.status = "error"
            db.commit()
        raise HTTPException(status_code=500, detail=f"Error uploading file: {str(e)}")
    finally:
        db.close()

@app.get("/api/mappings", response_model=List[MappingResponse])
def get_mappings():
    db = next(get_db())
    try:
        mappings = db.query(Mapping).order_by(Mapping.upload_date.desc()).all()
        return mappings
    finally:
        db.close()

@app.get("/api/mappings/{mapping_id}", response_model=MappingResponse)
def get_mapping(mapping_id: int):
    db = next(get_db())
    try:
        mapping = db.query(Mapping).filter(Mapping.id == mapping_id).first()
        if not mapping:
            raise HTTPException(status_code=404, detail="Mapping not found")
        return mapping
    finally:
        db.close()

@app.get("/api/mappings/{mapping_id}/fields", response_model=List[ParsedFieldResponse])
def get_parsed_fields(mapping_id: int):
    db = next(get_db())
    try:
        fields = db.query(ParsedField).filter(ParsedField.mapping_id == mapping_id).all()
        return fields
    finally:
        db.close()

@app.post("/api/excel/validate")
async def validate_excel_data(
    file: UploadFile = File(...),
    mapping_id: Optional[int] = Form(None),
    endpoint_id: int = Form(...),
    response_list_path: str = Form(...),
    match_key: str = Form(...),
    text_content_field: str = Form("textContent")
):
    db = next(get_db())
    try:
        if not file.filename:
            raise HTTPException(status_code=400, detail="No file provided")
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files allowed")

        content = await file.read()
        if len(content) == 0:
            raise HTTPException(status_code=400, detail="File is empty")

        import openpyxl
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        header_indexes = [idx for idx, name in enumerate(headers) if name]

        if not headers or match_key not in headers:
            raise HTTPException(status_code=400, detail=f"Match key '{match_key}' not found in Excel headers")

        endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == endpoint_id).first()
        if not endpoint:
            raise HTTPException(status_code=404, detail="Endpoint not found")

        url = f"{endpoint.base_url}{endpoint.path}"
        headers_dict = {}
        if endpoint.headers:
            headers_dict = json.loads(endpoint.headers)

        if endpoint.token_endpoint_id:
            token = get_or_fetch_token(endpoint.id, db)
            if token:
                headers_dict["Authorization"] = f"Bearer {token}"
        elif endpoint.auth_type == "bearer" and endpoint.auth_token:
            headers_dict["Authorization"] = f"Bearer {endpoint.auth_token}"
        elif endpoint.auth_type == "api_key" and endpoint.auth_token:
            headers_dict["X-API-Key"] = endpoint.auth_token

        request_body = None
        if endpoint.default_request_body:
            try:
                request_body = json.loads(endpoint.default_request_body)
            except Exception:
                request_body = None

        if endpoint.method.upper() == "GET":
            response = requests.get(url, headers=headers_dict, timeout=endpoint.timeout_ms / 1000)
        elif endpoint.method.upper() == "POST":
            response = requests.post(url, json=request_body or {}, headers=headers_dict, timeout=endpoint.timeout_ms / 1000)
        elif endpoint.method.upper() == "PUT":
            response = requests.put(url, json=request_body or {}, headers=headers_dict, timeout=endpoint.timeout_ms / 1000)
        elif endpoint.method.upper() == "DELETE":
            response = requests.delete(url, headers=headers_dict, timeout=endpoint.timeout_ms / 1000)
        else:
            response = requests.request(endpoint.method, url, json=request_body or {}, headers=headers_dict, timeout=endpoint.timeout_ms / 1000)

        response.raise_for_status()
        response_json = response.json()
        records = get_value_by_path(response_json, response_list_path)
        if not isinstance(records, list):
            raise HTTPException(status_code=400, detail="Response list path did not resolve to an array")

        results = []
        pass_rows = 0
        fail_rows = 0

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(value is None for value in row):
                continue

            row_data = {}
            for idx in header_indexes:
                header = headers[idx]
                if not header:
                    continue
                row_data[header] = row[idx] if idx < len(row) else None

            match_value = normalize_cell_value(row_data.get(match_key))
            matched_record = None
            for record in records:
                record_value = normalize_cell_value(record.get(match_key))
                if record_value == match_value:
                    matched_record = record
                    break

            if not matched_record:
                fail_rows += 1
                results.append({
                    "row_index": row_index,
                    "match_key": match_key,
                    "match_value": match_value,
                    "status": "fail",
                    "record_found": False,
                    "field_results": [],
                    "matched_record": None
                })
                continue

            field_results = compare_excel_row_to_record(row_data, matched_record, text_content_field)
            row_status = "pass" if all(item["status"] == "pass" for item in field_results) else "fail"
            if row_status == "pass":
                pass_rows += 1
            else:
                fail_rows += 1

            results.append({
                "row_index": row_index,
                "match_key": match_key,
                "match_value": match_value,
                "status": row_status,
                "record_found": True,
                "field_results": field_results,
                "matched_record": matched_record
            })

        total_rows = pass_rows + fail_rows
        return {
            "total_rows": total_rows,
            "pass_rows": pass_rows,
            "fail_rows": fail_rows,
            "response_status": response.status_code,
            "results": results
        }
    finally:
        db.close()

@app.put("/api/fields/{field_id}/approve")
def approve_field(field_id: int):
    db = next(get_db())
    try:
        field = db.query(ParsedField).filter(ParsedField.id == field_id).first()
        if not field:
            raise HTTPException(status_code=404, detail="Field not found")
        
        field.approved = True
        db.commit()
        db.refresh(field)
        return field
    finally:
        db.close()

@app.post("/api/mappings/{mapping_id}/generate-tests")
async def generate_tests(mapping_id: int):
    db = next(get_db())
    try:
        mapping = db.query(Mapping).filter(Mapping.id == mapping_id).first()
        if not mapping:
            raise HTTPException(status_code=404, detail="Mapping not found")
        
        scenarios_count = await generate_test_scenarios_ai(mapping_id, db)
        
        return {"scenarios_generated": scenarios_count}
    finally:
        db.close()

@app.get("/api/scenarios/all", response_model=List[TestScenarioResponse])
def get_all_scenarios():
    """Get all scenarios (for when no mapping is selected)"""
    db = next(get_db())
    try:
        scenarios = db.query(TestScenario).all()
        # Debug logging
        print(f"Total scenarios in database: {len(scenarios)}")
        for s in scenarios:
            print(f"  - ID: {s.id}, Name: {s.name}, Endpoint: {s.endpoint_id}, Mapping: {s.mapping_id}, Category: {s.category}, BusinessRule: {s.business_rule_id if hasattr(s, 'business_rule_id') else 'N/A'}")
        return scenarios
    finally:
        db.close()

@app.get("/api/scenarios/debug")
def debug_scenarios():
    """Debug endpoint to see raw scenario data"""
    db = next(get_db())
    try:
        scenarios = db.query(TestScenario).all()
        result = []
        for s in scenarios:
            result.append({
                "id": s.id,
                "name": s.name,
                "endpoint_id": s.endpoint_id,
                "mapping_id": s.mapping_id,
                "category": s.category,
                "business_rule_id": s.business_rule_id if hasattr(s, 'business_rule_id') else None
            })
        return {"total": len(result), "scenarios": result}
    finally:
        db.close()

@app.get("/api/mappings/{mapping_id}/scenarios", response_model=List[TestScenarioResponse])
def get_scenarios(mapping_id: int):
    db = next(get_db())
    try:
        scenarios = db.query(TestScenario).filter(TestScenario.mapping_id == mapping_id).all()
        return scenarios
    finally:
        db.close()

@app.post("/api/mappings/{mapping_id}/generate-scenarios")
def generate_scenarios(mapping_id: int):
    """Generate AI test scenarios for a mapping"""
    db = next(get_db())
    try:
        mapping = db.query(Mapping).filter(Mapping.id == mapping_id).first()
        if not mapping:
            raise HTTPException(status_code=404, detail="Mapping not found")
        
        scenarios_count = generate_test_scenarios_ai(mapping_id, db)
        
        return {"scenarios_generated": scenarios_count}
    finally:
        db.close()

@app.get("/api/mappings/{mapping_id}/scenarios/export")
def export_scenarios_to_excel(mapping_id: int):
    """Export scenarios to Excel with category, name, test type, and description"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    
    db = next(get_db())
    try:
        mapping = db.query(Mapping).filter(Mapping.id == mapping_id).first()
        if not mapping:
            raise HTTPException(status_code=404, detail="Mapping not found")
        
        scenarios = db.query(TestScenario).filter(TestScenario.mapping_id == mapping_id).all()
        
        if not scenarios:
            raise HTTPException(status_code=404, detail="No scenarios found for this mapping")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Scenarios"
        
        # Define headers
        headers = ["Test Name", "Category", "Test Type", "Description"]
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write scenario data
        for row_num, scenario in enumerate(scenarios, 2):
            ws.cell(row=row_num, column=1).value = scenario.name
            ws.cell(row=row_num, column=2).value = scenario.category or "N/A"
            
            # Determine test type based on scenario name/category
            test_type = "Functional"
            if "authentication" in scenario.name.lower() or "authorization" in scenario.name.lower():
                test_type = "Security"
            elif "boundary" in scenario.name.lower():
                test_type = "Boundary"
            elif "negative" in scenario.name.lower() or scenario.category == "negative":
                test_type = "Negative"
            elif "happy path" in scenario.name.lower() or scenario.category == "positive":
                test_type = "Positive"
            elif "data type" in scenario.name.lower():
                test_type = "Data Validation"
            elif "required field" in scenario.name.lower():
                test_type = "Field Validation"
            elif "business logic" in scenario.name.lower():
                test_type = "Business Logic"
            elif "header" in scenario.name.lower():
                test_type = "Header Validation"
            
            ws.cell(row=row_num, column=3).value = test_type
            ws.cell(row=row_num, column=4).value = scenario.description or "No description available"
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Return as downloadable file
        filename = f"test_scenarios_{mapping.filename.replace('.xlsx', '').replace('.xls', '')}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    finally:
        db.close()

async def execute_business_rule_scenario(scenario_id: int, db: Session):
    """Execute a business rule scenario and create test execution and validation results"""
    import time
    
    scenario = db.query(TestScenario).filter(TestScenario.id == scenario_id).first()
    if not scenario or not scenario.business_rule_id:
        raise HTTPException(status_code=404, detail="Business rule scenario not found")
    
    rule = db.query(CustomBusinessRule).filter(CustomBusinessRule.id == scenario.business_rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Business rule not found")
    
    endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == scenario.endpoint_id).first()
    if not endpoint:
        raise HTTPException(status_code=404, detail="Endpoint not found")
    
    start_time = time.time()
    
    # Call the API endpoint
    url = endpoint.base_url + endpoint.path
    headers = json.loads(endpoint.headers) if endpoint.headers else {}
    
    if endpoint.auth_type == "bearer" and endpoint.auth_token:
        headers["Authorization"] = f"Bearer {endpoint.auth_token}"
    
    # Make API call
    if endpoint.method.upper() == "GET":
        response = requests.get(url, headers=headers, timeout=endpoint.timeout_ms/1000)
    elif endpoint.method.upper() == "POST":
        body = json.loads(endpoint.default_request_body) if endpoint.default_request_body else {}
        response = requests.post(url, json=body, headers=headers, timeout=endpoint.timeout_ms/1000)
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported method: {endpoint.method}")
    
    response_time_ms = int((time.time() - start_time) * 1000)
    response_data = response.json()
    
    # Extract records from response
    records = []
    wrapper_fields = {}
    
    if isinstance(response_data, list):
        records = response_data
    elif isinstance(response_data, dict):
        found_array = False
        for key, value in response_data.items():
            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                records = value
                wrapper_fields = {k: v for k, v in response_data.items() if k != key}
                found_array = True
                break
        if not found_array:
            records = [response_data]
    else:
        records = [response_data]
    
    # Get schema fields
    schema_fields = []
    if wrapper_fields:
        for key in wrapper_fields.keys():
            if not isinstance(wrapper_fields[key], dict):
                schema_fields.append(key)
    if records and len(records) > 0:
        schema_fields.extend(list(records[0].keys()))
    
    # Interpret the rule using AI
    rule_interpretation = interpret_rule_with_ai(rule.rule_text, schema_fields)
    
    # Determine if wrapper or array rule
    validation_field = rule_interpretation.get("validation_field")
    condition_field = rule_interpretation.get("condition_field")
    wrapper_field_keys = list(wrapper_fields.keys()) if wrapper_fields else []
    is_wrapper_rule = (validation_field in wrapper_field_keys) or (condition_field in wrapper_field_keys)
    
    # Run validation
    validation_results = []
    pass_count = 0
    fail_count = 0
    
    if is_wrapper_rule:
        wrapper_record = {**wrapper_fields, **(records[0] if records else {})}
        validation_result = apply_validation_rule(wrapper_record, rule_interpretation, 0)
        validation_result["record_data"] = wrapper_record  # Add the actual record data
        validation_results.append(validation_result)
        if validation_result["result"] == "PASS":
            pass_count += 1
        else:
            fail_count += 1
    else:
        for idx, record in enumerate(records, start=1):
            enriched_record = {**wrapper_fields, **record}
            validation_result = apply_validation_rule(enriched_record, rule_interpretation, idx)
            validation_result["record_data"] = record  # Add the actual record data (without wrapper fields for clarity)
            validation_results.append(validation_result)
            if validation_result["result"] == "PASS":
                pass_count += 1
            else:
                fail_count += 1
    
    # Create test execution
    execution = TestExecution(
        scenario_id=scenario_id,
        status="completed",
        pass_count=pass_count,
        fail_count=fail_count,
        total_response_time_ms=response_time_ms,
        request_body=endpoint.default_request_body,
        response_body=json.dumps(response_data),
        expected_response=rule.rule_text
    )
    db.add(execution)
    db.commit()
    db.refresh(execution)
    
    # Create validation results
    for val_result in validation_results:
        validation = ValidationResult(
            scenario_id=scenario_id,
            execution_id=execution.id,
            field_name=val_result["field_name"] or "business_rule",
            expected=val_result["expected_value"],
            actual=val_result["actual_value"],
            status="pass" if val_result["result"] == "PASS" else "fail",
            validation_type="BUSINESS_RULE",
            response_time_ms=response_time_ms,
            status_code=response.status_code
        )
        db.add(validation)
    
    db.commit()
    
    # Update scenario status
    scenario.status = "completed" if fail_count == 0 else "failed"
    db.commit()
    
    # Format results to match frontend expectations
    formatted_results = []
    for val_result in validation_results:
        formatted_results.append({
            "field_name": val_result.get("field_name") or "business_rule",
            "validation_type": "BUSINESS_RULE",
            "expected": val_result.get("expected_value"),
            "actual": val_result.get("actual_value"),
            "status": "pass" if val_result.get("result") == "PASS" else "fail",
            "error_message": val_result.get("error_message"),
            "record_data": val_result.get("record_data"),  # Include the specific record that was validated
            "record_number": val_result.get("record_number", 0)
        })
    
    return {
        "execution_id": execution.id,
        "status": execution.status,
        "pass_count": pass_count,
        "fail_count": fail_count,
        "response_time_ms": response_time_ms,
        "results": formatted_results  # Changed from validation_results to results
    }

@app.post("/api/scenarios/{scenario_id}/execute")
async def execute_scenario(scenario_id: int, request_body: dict = None):
    db = next(get_db())
    try:
        print(f"Executing scenario ID: {scenario_id}")
        scenario = db.query(TestScenario).filter(TestScenario.id == scenario_id).first()
        if not scenario:
            raise HTTPException(status_code=404, detail="Scenario not found")
        
        print(f"Scenario found: {scenario.name}, Category: {scenario.category}, Business Rule ID: {scenario.business_rule_id if hasattr(scenario, 'business_rule_id') else 'N/A'}")
        
        # Check if this is a business rule scenario
        if hasattr(scenario, 'business_rule_id') and scenario.business_rule_id:
            # Execute business rule validation
            print(f"Executing business rule scenario...")
            result = await execute_business_rule_scenario(scenario_id, db)
            print(f"Business rule execution completed successfully")
        else:
            # Execute standard validation
            print(f"Executing standard validation...")
            result = execute_validation(scenario_id, db, request_body)
            print(f"Standard validation completed successfully")
        
        return result
    except Exception as e:
        import traceback
        print(f"✗ ERROR executing scenario {scenario_id}: {str(e)}")
        print(f"Full traceback:")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error executing scenario: {str(e)}")
    finally:
        db.close()

@app.post("/api/scenarios/{scenario_id}/execute-schema")
async def execute_scenario_schema_only(scenario_id: int):
    """Execute schema-only validation (no type checking, only structure and field presence)"""
    from schema_endpoint import execute_schema_validation_endpoint
    
    db = next(get_db())
    try:
        result = execute_schema_validation_endpoint(scenario_id, db)
        return result
    finally:
        db.close()

@app.get("/api/scenarios/{scenario_id}/results", response_model=List[ValidationResultResponse])
def get_results(scenario_id: int):
    db = next(get_db())
    try:
        results = db.query(ValidationResult).filter(
            ValidationResult.scenario_id == scenario_id
        ).order_by(ValidationResult.timestamp.desc()).all()
        return results
    finally:
        db.close()

# AI Root Cause Analysis Endpoints
@app.post("/api/failures/{validation_result_id}/analyze")
def analyze_failure(validation_result_id: int):
    """Analyze a single validation failure with AI and return root cause analysis"""
    db = next(get_db())
    try:
        analysis = analyze_failure_with_ai(validation_result_id, db)
        return analysis
    finally:
        db.close()

@app.post("/api/scenarios/{scenario_id}/analyze-all-failures")
def analyze_all_failures(scenario_id: int):
    """Analyze all failures from a scenario execution in batch"""
    import time
    
    db = next(get_db())
    try:
        # Get all failed validations for this scenario
        failed_validations = db.query(ValidationResult).filter(
            ValidationResult.scenario_id == scenario_id,
            ValidationResult.status == 'fail'
        ).all()
        
        if not failed_validations:
            return {
                'total_failures': 0,
                'analyses': [],
                'message': 'No failures found for this scenario'
            }
        
        results = []
        
        for validation in failed_validations:
            # Skip if already analyzed
            if validation.root_cause and validation.root_cause_category:
                results.append({
                    'validation_result_id': validation.id,
                    'field': validation.field_name,
                    'root_cause': validation.root_cause,
                    'category': validation.root_cause_category,
                    'suggested_action': validation.suggested_action,
                    'confidence': validation.ai_confidence,
                    'cached': True
                })
                continue
            
            # Analyze this failure
            analysis = analyze_failure_with_ai(validation.id, db)
            
            if 'error' not in analysis:
                results.append({
                    'validation_result_id': validation.id,
                    'field': validation.field_name,
                    'root_cause': analysis.get('explanation', ''),
                    'category': analysis.get('root_cause_category', 'UNKNOWN'),
                    'suggested_action': analysis.get('suggested_action', ''),
                    'confidence': analysis.get('confidence', 0),
                    'cached': analysis.get('from_cache', False)
                })
            else:
                results.append({
                    'validation_result_id': validation.id,
                    'field': validation.field_name,
                    'error': analysis['error']
                })
            
            # Small delay to avoid rate limits
            time.sleep(0.5)
        
        return {
            'total_failures': len(failed_validations),
            'analyses': results
        }
    finally:
        db.close()

@app.get("/api/scenarios/{scenario_id}/insights")
def get_ai_insights(scenario_id: int):
    """Generate AI-powered insights by analyzing patterns across all failures"""
    import json
    
    db = next(get_db())
    try:
        # Get all failures
        failures = db.query(ValidationResult).filter(
            ValidationResult.scenario_id == scenario_id,
            ValidationResult.status == 'fail'
        ).all()
        
        if not failures:
            return {
                'insights': [],
                'message': 'No failures to analyze'
            }
        
        # Group by root cause category
        by_category = {}
        for f in failures:
            cat = f.root_cause_category or 'Unknown'
            if cat not in by_category:
                by_category[cat] = []
            by_category[cat].append(f)
        
        # Build prompt for insights generation
        prompt = f"""Analyze these API validation failures and identify the top 3-5 actionable insights.

Total failures: {len(failures)}

Failures by category:
{json.dumps({k: len(v) for k, v in by_category.items()}, indent=2)}

Sample failures:
{json.dumps([{
    'field': f.field_name,
    'category': f.root_cause_category,
    'root_cause': f.root_cause[:100] if f.root_cause else 'Not analyzed'
} for f in failures[:10]], indent=2)}

Generate 3-5 insights that:
1. Identify PATTERNS across multiple failures
2. Prioritize by SEVERITY (critical/warning/info)
3. Provide ACTIONABLE next steps
4. Are SPECIFIC, not generic

Return ONLY valid JSON array (no markdown, no code blocks):
[
  {{
    "severity": "critical|warning|info",
    "title": "Brief title (5-8 words)",
    "description": "1-2 sentence explanation",
    "affected_fields": 5,
    "suggested_action": "Specific next step"
  }}
]

Focus on insights like:
- "60% of failures are in address fields — vendor data incomplete"
- "Status field inconsistent with endDate across all records"
- "API returning stale cache — lastUpdateDate is 3 days old"
"""
        
        # Call Azure OpenAI
        azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
        api_key = os.getenv("AZURE_OPENAI_API_KEY")
        api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")
        deployment_name = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "gpt-4")
        
        if not api_key or not azure_endpoint:
            return {"error": "Azure OpenAI not configured"}
        
        client = AzureOpenAI(
            api_key=api_key,
            api_version=api_version,
            azure_endpoint=azure_endpoint
        )
        
        response = client.chat.completions.create(
            model=deployment_name,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=1000,
            temperature=0.3
        )
        
        # Parse AI response
        response_text = response.choices[0].message.content.strip()
        
        # Clean up response
        if response_text.startswith('```'):
            response_text = response_text.split('```')[1]
            if response_text.startswith('json'):
                response_text = response_text[4:]
            response_text = response_text.strip()
        
        insights = json.loads(response_text)
        
        return {
            'insights': insights,
            'total_failures': len(failures),
            'categories': {k: len(v) for k, v in by_category.items()}
        }
        
    except Exception as e:
        print(f"Error generating insights: {e}")
        import traceback
        traceback.print_exc()
        return {"error": f"Failed to generate insights: {str(e)}"}
    finally:
        db.close()

# Scenario Update Endpoints
@app.post("/api/scenarios/create", response_model=TestScenarioResponse)
def create_custom_scenario(scenario_data: dict):
    """Create a custom test scenario without Excel upload"""
    db = next(get_db())
    try:
        # Create a dummy mapping if none exists (for custom scenarios)
        mapping_id = scenario_data.get('mapping_id')
        if not mapping_id:
            # Create or get a default "Custom Scenarios" mapping
            custom_mapping = db.query(Mapping).filter(Mapping.filename == "Custom Scenarios").first()
            if not custom_mapping:
                custom_mapping = Mapping(
                    filename="Custom Scenarios",
                    status="completed",
                    parsed_count=0
                )
                db.add(custom_mapping)
                db.commit()
                db.refresh(custom_mapping)
            mapping_id = custom_mapping.id
        
        # Create the scenario
        new_scenario = TestScenario(
            mapping_id=mapping_id,
            endpoint_id=scenario_data.get('endpoint_id'),
            name=scenario_data.get('name', 'Untitled Scenario'),
            category=scenario_data.get('category', 'custom'),
            status='pending',
            request_body=scenario_data.get('request_body'),
            expected_response=scenario_data.get('expected_response'),
            json_schema=scenario_data.get('json_schema')
        )
        
        db.add(new_scenario)
        db.commit()
        db.refresh(new_scenario)
        
        return new_scenario
    finally:
        db.close()

@app.put("/api/scenarios/{scenario_id}", response_model=TestScenarioResponse)
def update_scenario(scenario_id: int, update: TestScenarioUpdate):
    db = next(get_db())
    try:
        scenario = db.query(TestScenario).filter(TestScenario.id == scenario_id).first()
        if not scenario:
            raise HTTPException(status_code=404, detail="Scenario not found")
        
        if update.endpoint_id is not None:
            scenario.endpoint_id = update.endpoint_id
        if update.request_body is not None:
            scenario.request_body = update.request_body
        if update.expected_response is not None:
            scenario.expected_response = update.expected_response
        if update.json_schema is not None:
            scenario.json_schema = update.json_schema
        
        db.commit()
        db.refresh(scenario)
        return scenario
    finally:
        db.close()

@app.delete("/api/scenarios/delete-all")
def delete_all_scenarios():
    """Delete all test scenarios and all related data"""
    db = next(get_db())
    try:
        # Get count before deletion
        scenario_count = db.query(TestScenario).count()
        
        if scenario_count == 0:
            return {"message": "No scenarios to delete", "deleted_count": 0}
        
        # Delete all validation results
        db.query(ValidationResult).delete(synchronize_session=False)
        
        # Delete all test executions
        db.query(TestExecution).delete(synchronize_session=False)
        
        # Delete all root cause cache entries
        db.query(RootCauseCache).delete(synchronize_session=False)
        
        # Delete all scenarios
        db.query(TestScenario).delete(synchronize_session=False)
        
        db.commit()
        
        return {"message": f"Successfully deleted {scenario_count} scenarios", "deleted_count": scenario_count}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete scenarios: {str(e)}")
    finally:
        db.close()

@app.delete("/api/scenarios/{scenario_id}")
def delete_scenario(scenario_id: int):
    """Delete a test scenario and all related data"""
    db = next(get_db())
    try:
        scenario = db.query(TestScenario).filter(TestScenario.id == scenario_id).first()
        if not scenario:
            raise HTTPException(status_code=404, detail="Scenario not found")
        
        # Delete related validation results
        db.query(ValidationResult).filter(ValidationResult.scenario_id == scenario_id).delete()
        
        # Delete related executions
        db.query(TestExecution).filter(TestExecution.scenario_id == scenario_id).delete()
        
        # Delete the scenario
        db.delete(scenario)
        db.commit()
        
        return {"message": "Scenario deleted successfully", "scenario_id": scenario_id}
    finally:
        db.close()


# Test History Endpoints
@app.get("/api/scenarios/{scenario_id}/executions", response_model=List[TestExecutionResponse])
def get_execution_history(scenario_id: int, limit: int = 10):
    db = next(get_db())
    try:
        executions = db.query(TestExecution).filter(
            TestExecution.scenario_id == scenario_id
        ).order_by(TestExecution.execution_date.desc()).limit(limit).all()
        return executions
    finally:
        db.close()

@app.get("/api/executions/{execution_id}", response_model=TestExecutionResponse)
def get_execution(execution_id: int):
    db = next(get_db())
    try:
        execution = db.query(TestExecution).filter(TestExecution.id == execution_id).first()
        if not execution:
            raise HTTPException(status_code=404, detail="Execution not found")
        return execution
    finally:
        db.close()

@app.get("/api/executions/{execution_id}/results", response_model=List[ValidationResultResponse])
def get_execution_results(execution_id: int):
    db = next(get_db())
    try:
        results = db.query(ValidationResult).filter(
            ValidationResult.execution_id == execution_id
        ).order_by(ValidationResult.timestamp).all()
        return results
    finally:
        db.close()

# Export Endpoints
@app.get("/api/scenarios/{scenario_id}/export/csv")
def export_results_csv(scenario_id: int, execution_id: Optional[int] = None):
    import csv
    from io import StringIO
    from fastapi.responses import StreamingResponse
    
    db = next(get_db())
    try:
        query = db.query(ValidationResult).filter(ValidationResult.scenario_id == scenario_id)
        if execution_id:
            query = query.filter(ValidationResult.execution_id == execution_id)
        results = query.order_by(ValidationResult.timestamp.desc()).all()
        
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Field Name', 'Expected', 'Actual', 'Status', 'Validation Type', 'Root Cause', 'Response Time (ms)', 'Status Code', 'Timestamp'])
        
        for result in results:
            writer.writerow([
                result.field_name,
                result.expected,
                result.actual,
                result.status,
                result.validation_type,
                result.root_cause,
                result.response_time_ms,
                result.status_code,
                result.timestamp.isoformat()
            ])
        
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=scenario_{scenario_id}_results.csv"}
        )
    finally:
        db.close()

@app.get("/api/scenarios/{scenario_id}/export/json")
def export_results_json(scenario_id: int, execution_id: Optional[int] = None):
    import json
    from fastapi.responses import StreamingResponse
    
    db = next(get_db())
    try:
        query = db.query(ValidationResult).filter(ValidationResult.scenario_id == scenario_id)
        if execution_id:
            query = query.filter(ValidationResult.execution_id == execution_id)
        results = query.order_by(ValidationResult.timestamp.desc()).all()
        
        data = [{
            "field_name": r.field_name,
            "expected": r.expected,
            "actual": r.actual,
            "status": r.status,
            "validation_type": r.validation_type,
            "root_cause": r.root_cause,
            "response_time_ms": r.response_time_ms,
            "status_code": r.status_code,
            "timestamp": r.timestamp.isoformat()
        } for r in results]
        
        json_str = json.dumps(data, indent=2)
        return StreamingResponse(
            iter([json_str]),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename=scenario_{scenario_id}_results.json"}
        )
    finally:
        db.close()

# Environment Management Endpoints
@app.post("/api/environments", response_model=EnvironmentResponse)
def create_environment(env: EnvironmentCreate):
    db = next(get_db())
    try:
        new_env = Environment(
            name=env.name,
            description=env.description,
            variables=env.variables,
            is_active=False
        )
        db.add(new_env)
        db.commit()
        db.refresh(new_env)
        return new_env
    finally:
        db.close()

@app.get("/api/environments", response_model=List[EnvironmentResponse])
def get_environments():
    db = next(get_db())
    try:
        environments = db.query(Environment).all()
        return environments
    finally:
        db.close()

@app.get("/api/environments/{env_id}", response_model=EnvironmentResponse)
def get_environment(env_id: int):
    db = next(get_db())
    try:
        env = db.query(Environment).filter(Environment.id == env_id).first()
        if not env:
            raise HTTPException(status_code=404, detail="Environment not found")
        return env
    finally:
        db.close()

@app.put("/api/environments/{env_id}/activate")
def activate_environment(env_id: int):
    db = next(get_db())
    try:
        db.query(Environment).update({"is_active": False})
        env = db.query(Environment).filter(Environment.id == env_id).first()
        if not env:
            raise HTTPException(status_code=404, detail="Environment not found")
        env.is_active = True
        db.commit()
        return {"message": f"Environment '{env.name}' activated"}
    finally:
        db.close()

@app.delete("/api/environments/{env_id}")
def delete_environment(env_id: int):
    db = next(get_db())
    try:
        env = db.query(Environment).filter(Environment.id == env_id).first()
        if not env:
            raise HTTPException(status_code=404, detail="Environment not found")
        db.delete(env)
        db.commit()
        return {"message": "Environment deleted"}
    finally:
        db.close()

# API Endpoint Management
@app.post("/api/endpoints", response_model=APIEndpointResponse)
def create_endpoint(endpoint: APIEndpointCreate):
    db = next(get_db())
    try:
        new_endpoint = APIEndpoint(
            name=endpoint.name,
            environment_id=endpoint.environment_id,
            base_url=endpoint.base_url,
            method=endpoint.method,
            path=endpoint.path,
            auth_type=endpoint.auth_type,
            auth_token=endpoint.auth_token,
            headers=endpoint.headers,
            timeout_ms=endpoint.timeout_ms,
            expected_status=endpoint.expected_status,
            max_response_time_ms=endpoint.max_response_time_ms
        )
        db.add(new_endpoint)
        db.commit()
        db.refresh(new_endpoint)
        return new_endpoint
    finally:
        db.close()

@app.get("/api/endpoints", response_model=List[APIEndpointResponse])
def get_endpoints(environment_id: Optional[int] = None):
    db = next(get_db())
    try:
        query = db.query(APIEndpoint)
        if environment_id:
            query = query.filter(APIEndpoint.environment_id == environment_id)
        endpoints = query.all()
        return endpoints
    finally:
        db.close()

@app.get("/api/endpoints/{endpoint_id}", response_model=APIEndpointResponse)
def get_endpoint(endpoint_id: int):
    db = next(get_db())
    try:
        endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == endpoint_id).first()
        if not endpoint:
            raise HTTPException(status_code=404, detail="Endpoint not found")
        return endpoint
    finally:
        db.close()

@app.put("/api/endpoints/{endpoint_id}", response_model=APIEndpointResponse)
def update_endpoint(endpoint_id: int, endpoint: APIEndpointCreate):
    db = next(get_db())
    try:
        db_endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == endpoint_id).first()
        if not db_endpoint:
            raise HTTPException(status_code=404, detail="Endpoint not found")
        
        for key, value in endpoint.model_dump().items():
            setattr(db_endpoint, key, value)
        
        db.commit()
        db.refresh(db_endpoint)
        return db_endpoint
    finally:
        db.close()

@app.delete("/api/endpoints/{endpoint_id}")
def delete_endpoint(endpoint_id: int):
    db = next(get_db())
    try:
        endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == endpoint_id).first()
        if not endpoint:
            raise HTTPException(status_code=404, detail="Endpoint not found")
        db.delete(endpoint)
        db.commit()
        return {"message": "Endpoint deleted"}
    finally:
        db.close()

@app.post("/api/endpoints/test")
async def test_endpoint_connection(endpoint_data: dict):
    """Test an endpoint connection without saving it"""
    import time
    
    try:
        url = f"{endpoint_data.get('base_url', '')}{endpoint_data.get('path', '')}"
        method = endpoint_data.get('method', 'GET')
        headers = {}
        
        # Parse headers if provided
        if endpoint_data.get('headers'):
            try:
                headers = json.loads(endpoint_data['headers'])
            except json.JSONDecodeError:
                return {
                    "success": False,
                    "message": "Invalid JSON in headers",
                    "error": "Headers must be valid JSON"
                }
        
        # Add authentication headers
        auth_type = endpoint_data.get('auth_type', '')
        auth_token = endpoint_data.get('auth_token', '')
        
        if auth_type == 'bearer' and auth_token:
            headers['Authorization'] = f'Bearer {auth_token}'
        elif auth_type == 'api_key' and auth_token:
            headers['X-API-Key'] = auth_token
        
        # Parse request body if provided (for POST/PUT/PATCH)
        request_body = None
        body_data = None
        content_type = headers.get('Content-Type', '').lower()
        
        if endpoint_data.get('default_request_body'):
            body_str = endpoint_data['default_request_body']
            
            # Check if it's form-urlencoded format
            if 'application/x-www-form-urlencoded' in content_type:
                # Keep as string for form data
                body_data = body_str
            else:
                # Try to parse as JSON
                try:
                    request_body = json.loads(body_str)
                except json.JSONDecodeError:
                    # If not valid JSON, treat as raw string
                    body_data = body_str
        
        # Make the request
        timeout_ms = endpoint_data.get('timeout_ms', 5000)
        timeout_sec = timeout_ms / 1000
        
        start_time = time.time()
        
        # Send request with appropriate body format
        if body_data:
            # Send as raw data (for form-urlencoded)
            response = requests.request(
                method=method,
                url=url,
                headers=headers,
                data=body_data,
                timeout=timeout_sec
            )
        elif request_body:
            # Send as JSON
            response = requests.request(
                method=method,
                url=url,
                headers=headers,
                json=request_body,
                timeout=timeout_sec
            )
        else:
            # No body
            response = requests.request(
                method=method,
                url=url,
                headers=headers,
                timeout=timeout_sec
            )
        
        end_time = time.time()
        response_time = int((end_time - start_time) * 1000)
        
        # Parse response body
        try:
            response_body = response.json()
        except:
            response_body = response.text
        
        # Check if status matches expected
        expected_status = endpoint_data.get('expected_status', 200)
        success = response.status_code == expected_status
        
        # Check SLA
        max_response_time = endpoint_data.get('max_response_time_ms', 2000)
        within_sla = response_time <= max_response_time
        
        return {
            "success": success,
            "status": response.status_code,
            "responseTime": response_time,
            "message": f"✅ Connection successful! Status: {response.status_code}, Time: {response_time}ms" if success else f"⚠️ Status mismatch: Expected {expected_status}, Got {response.status_code}",
            "responseBody": response_body,
            "withinSLA": within_sla
        }
        
    except requests.exceptions.Timeout:
        return {
            "success": False,
            "message": f"❌ Connection timeout after {timeout_ms}ms",
            "error": "Request timed out"
        }
    except requests.exceptions.ConnectionError as e:
        return {
            "success": False,
            "message": f"❌ Connection failed: Unable to reach endpoint",
            "error": str(e)
        }
    except Exception as e:
        return {
            "success": False,
            "message": f"❌ Connection error: {str(e)}",
            "error": str(e)
        }

@app.get("/api/dashboard/stats")
def get_dashboard_stats():
    """Get comprehensive dashboard statistics for all test executions"""
    db = next(get_db())
    try:
        # Get all executions
        executions = db.query(TestExecution).all()
        
        total_executions = len(executions)
        total_passed = sum(1 for ex in executions if ex.fail_count == 0 and ex.status == 'completed')
        total_failed = sum(1 for ex in executions if ex.fail_count > 0 and ex.status == 'completed')
        success_rate = round((total_passed / total_executions * 100) if total_executions > 0 else 0, 1)
        
        # Get endpoint statistics
        endpoints = db.query(APIEndpoint).all()
        endpoint_stats = []
        
        for endpoint in endpoints:
            # Get all scenarios for this endpoint
            scenarios = db.query(TestScenario).filter(TestScenario.endpoint_id == endpoint.id).all()
            scenario_ids = [s.id for s in scenarios]
            
            if scenario_ids:
                # Get executions for these scenarios
                endpoint_executions = db.query(TestExecution).filter(
                    TestExecution.scenario_id.in_(scenario_ids)
                ).all()
                
                total_runs = len(endpoint_executions)
                passed = sum(1 for ex in endpoint_executions if ex.fail_count == 0 and ex.status == 'completed')
                failed = sum(1 for ex in endpoint_executions if ex.fail_count > 0 and ex.status == 'completed')
                
                # Calculate average response time
                response_times = [ex.total_response_time_ms for ex in endpoint_executions if ex.total_response_time_ms]
                avg_response_time = round(sum(response_times) / len(response_times)) if response_times else 0
                
                endpoint_stats.append({
                    'name': endpoint.name,
                    'method': endpoint.method,
                    'path': endpoint.path,
                    'total_runs': total_runs,
                    'passed': passed,
                    'failed': failed,
                    'success_rate': round((passed / total_runs * 100) if total_runs > 0 else 0, 1),
                    'avg_response_time': avg_response_time
                })
        
        # Get scenario statistics
        scenarios = db.query(TestScenario).all()
        scenario_stats = []
        
        for scenario in scenarios:
            # Get endpoint name
            endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == scenario.endpoint_id).first()
            endpoint_name = endpoint.name if endpoint else 'Unknown'
            
            # Get executions for this scenario
            scenario_executions = db.query(TestExecution).filter(
                TestExecution.scenario_id == scenario.id
            ).order_by(TestExecution.execution_date.desc()).all()
            
            total_runs = len(scenario_executions)
            passed = sum(1 for ex in scenario_executions if ex.fail_count == 0 and ex.status == 'completed')
            failed = sum(1 for ex in scenario_executions if ex.fail_count > 0 and ex.status == 'completed')
            
            last_execution = scenario_executions[0] if scenario_executions else None
            
            scenario_stats.append({
                'name': scenario.name,
                'description': scenario.description,
                'endpoint_name': endpoint_name,
                'total_runs': total_runs,
                'passed': passed,
                'failed': failed,
                'last_run': last_execution.execution_date.isoformat() if last_execution else None,
                'last_status': last_execution.status if last_execution else None,
                'last_fail_count': last_execution.fail_count if last_execution else 0
            })
        
        # Get recent executions (last 10)
        recent_executions = db.query(TestExecution).order_by(
            TestExecution.execution_date.desc()
        ).limit(10).all()
        
        recent_exec_data = []
        for execution in recent_executions:
            scenario = db.query(TestScenario).filter(TestScenario.id == execution.scenario_id).first()
            endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == scenario.endpoint_id).first() if scenario else None
            
            recent_exec_data.append({
                'scenario_name': scenario.name if scenario else 'Unknown',
                'endpoint_name': endpoint.name if endpoint else 'Unknown',
                'pass_count': execution.pass_count,
                'fail_count': execution.fail_count,
                'response_time': execution.total_response_time_ms or 0,
                'timestamp': execution.execution_date.isoformat()
            })
        
        return {
            'totalExecutions': total_executions,
            'totalPassed': total_passed,
            'totalFailed': total_failed,
            'successRate': success_rate,
            'endpointStats': endpoint_stats,
            'scenarioStats': scenario_stats,
            'recentExecutions': recent_exec_data
        }
        
    finally:
        db.close()

@app.get("/api/scenarios/export-excel")
def export_test_scenarios_to_excel(endpoint_id: Optional[int] = None, mapping_id: Optional[int] = None):
    """Export test scenarios with their latest execution results based on filters"""
    db = next(get_db())
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import StreamingResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Scenarios"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Set headers
        headers = ['Scenario Name', 'Category', 'Endpoint', 'Description', 'Status', 'Last Execution', 'Pass Count', 'Fail Count', 'Response Time (ms)']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Query scenarios based on filters
        query = db.query(TestScenario)
        
        if mapping_id:
            query = query.filter(TestScenario.mapping_id == mapping_id)
        elif endpoint_id:
            query = query.filter(TestScenario.endpoint_id == endpoint_id)
        # If no filters, get all scenarios
        
        scenarios = query.all()
        
        # Populate summary data
        row = 2
        for scenario in scenarios:
            # Get endpoint name
            endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == scenario.endpoint_id).first() if scenario.endpoint_id else None
            endpoint_name = endpoint.name if endpoint else 'N/A'
            
            # Get latest execution
            latest_execution = db.query(TestExecution).filter(
                TestExecution.scenario_id == scenario.id
            ).order_by(TestExecution.execution_date.desc()).first()
            
            ws.cell(row=row, column=1).value = scenario.name
            ws.cell(row=row, column=2).value = scenario.category or 'N/A'
            ws.cell(row=row, column=3).value = endpoint_name
            ws.cell(row=row, column=4).value = scenario.description or 'N/A'
            ws.cell(row=row, column=5).value = scenario.status or 'pending'
            
            if latest_execution:
                ws.cell(row=row, column=6).value = latest_execution.execution_date.strftime('%Y-%m-%d %H:%M:%S') if latest_execution.execution_date else 'N/A'
                ws.cell(row=row, column=7).value = latest_execution.pass_count or 0
                ws.cell(row=row, column=8).value = latest_execution.fail_count or 0
                ws.cell(row=row, column=9).value = latest_execution.total_response_time_ms or 0
            else:
                ws.cell(row=row, column=6).value = 'Never executed'
                ws.cell(row=row, column=7).value = 0
                ws.cell(row=row, column=8).value = 0
                ws.cell(row=row, column=9).value = 0
            
            row += 1
        
        # Create a second sheet for failure details
        ws_failures = wb.create_sheet("Failure Details")
        
        # Failure sheet headers
        failure_headers = ['Scenario Name', 'Field', 'Validation Type', 'Expected', 'Actual', 'Status', 'Error Message', 'Record Number', 'Record Data']
        for col_num, header in enumerate(failure_headers, 1):
            cell = ws_failures.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Populate failure details
        failure_row = 2
        for scenario in scenarios:
            # Get latest execution
            latest_execution = db.query(TestExecution).filter(
                TestExecution.scenario_id == scenario.id
            ).order_by(TestExecution.execution_date.desc()).first()
            
            if latest_execution and latest_execution.fail_count > 0:
                # For business rule scenarios, get the detailed failure records from BusinessRuleValidation
                if scenario.category == 'business_rule':
                    # Get business rule validations that failed
                    business_validations = db.query(BusinessRuleValidation).filter(
                        BusinessRuleValidation.endpoint_id == scenario.endpoint_id,
                        BusinessRuleValidation.result == 'FAIL'
                    ).order_by(BusinessRuleValidation.execution_date.desc()).all()
                    
                    # Parse the response to extract individual records
                    try:
                        import json
                        response_data = json.loads(latest_execution.response_body) if isinstance(latest_execution.response_body, str) else latest_execution.response_body
                        
                        # Extract the array of records from response
                        records = []
                        if isinstance(response_data, list):
                            records = response_data
                        elif isinstance(response_data, dict):
                            # Look for common array keys
                            for key in ['termsAndConditions', 'data', 'results', 'items', 'records']:
                                if key in response_data and isinstance(response_data[key], list):
                                    records = response_data[key]
                                    break
                        
                        # Match business validations with their specific records
                        for bv in business_validations:
                            ws_failures.cell(row=failure_row, column=1).value = scenario.name
                            ws_failures.cell(row=failure_row, column=2).value = bv.field_name or 'N/A'
                            ws_failures.cell(row=failure_row, column=3).value = 'BUSINESS_RULE'
                            ws_failures.cell(row=failure_row, column=4).value = bv.expected_value or 'N/A'
                            ws_failures.cell(row=failure_row, column=5).value = bv.actual_value or 'N/A'
                            ws_failures.cell(row=failure_row, column=6).value = bv.result or 'FAIL'
                            ws_failures.cell(row=failure_row, column=7).value = bv.error_message or 'N/A'
                            ws_failures.cell(row=failure_row, column=8).value = bv.record_number or 'N/A'
                            
                            # Get the specific record that failed
                            if records and bv.record_number and bv.record_number > 0 and bv.record_number <= len(records):
                                specific_record = records[bv.record_number - 1]  # record_number is 1-indexed
                                ws_failures.cell(row=failure_row, column=9).value = json.dumps(specific_record, indent=2)
                            else:
                                ws_failures.cell(row=failure_row, column=9).value = 'N/A'
                            
                            failure_row += 1
                    except Exception as e:
                        print(f"Error extracting record data: {str(e)}")
                        ws_failures.cell(row=failure_row, column=1).value = scenario.name
                        ws_failures.cell(row=failure_row, column=9).value = f"Error: {str(e)}"
                        failure_row += 1
                else:
                    # For non-business-rule scenarios, use ValidationResult table
                    validations = db.query(ValidationResult).filter(
                        ValidationResult.execution_id == latest_execution.id,
                        ValidationResult.status == 'fail'
                    ).all()
                    
                    for validation in validations:
                        ws_failures.cell(row=failure_row, column=1).value = scenario.name
                        ws_failures.cell(row=failure_row, column=2).value = validation.field_name or 'N/A'
                        ws_failures.cell(row=failure_row, column=3).value = validation.validation_type or 'FIELD'
                        ws_failures.cell(row=failure_row, column=4).value = validation.expected or 'N/A'
                        ws_failures.cell(row=failure_row, column=5).value = validation.actual or 'N/A'
                        ws_failures.cell(row=failure_row, column=6).value = validation.status or 'fail'
                        ws_failures.cell(row=failure_row, column=7).value = validation.root_cause or 'N/A'
                        ws_failures.cell(row=failure_row, column=8).value = 'N/A'
                        ws_failures.cell(row=failure_row, column=9).value = 'N/A'
                        
                        failure_row += 1
        
        # Auto-size columns for failures sheet
        for col in ws_failures.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    cell_value = str(cell.value)
                    # Limit very long JSON strings for column width calculation
                    if len(cell_value) > 100:
                        max_length = 100
                    elif len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws_failures.column_dimensions[column].width = adjusted_width
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"test_scenarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    finally:
        db.close()

@app.get("/api/dashboard/export-excel")
def export_dashboard_to_excel():
    """Export execution results to Excel with multiple sheets: Overall + per-endpoint"""
    db = next(get_db())
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import PieChart, Reference
        from openpyxl.chart.series import DataPoint
        from fastapi.responses import StreamingResponse
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Get all executions
        executions = db.query(TestExecution).all()
        total_executions = len(executions)
        total_passed = sum(1 for ex in executions if ex.fail_count == 0 and ex.status == 'completed')
        total_failed = sum(1 for ex in executions if ex.fail_count > 0 and ex.status == 'completed')
        success_rate = round((total_passed / total_executions * 100) if total_executions > 0 else 0, 1)
        
        # Sheet 1: Overall Summary
        ws_overall = wb.create_sheet("Overall Summary")
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Overall stats
        ws_overall['A1'] = 'Metric'
        ws_overall['B1'] = 'Value'
        ws_overall['A1'].fill = header_fill
        ws_overall['A1'].font = header_font
        ws_overall['B1'].fill = header_fill
        ws_overall['B1'].font = header_font
        
        ws_overall['A2'] = 'Total Executions'
        ws_overall['B2'] = total_executions
        ws_overall['A3'] = 'Total Passed'
        ws_overall['B3'] = total_passed
        ws_overall['A4'] = 'Total Failed'
        ws_overall['B4'] = total_failed
        ws_overall['A5'] = 'Success Rate (%)'
        ws_overall['B5'] = success_rate
        
        # Add pie chart data in columns D-E for the chart
        ws_overall['D1'] = 'Status'
        ws_overall['E1'] = 'Count'
        ws_overall['D1'].fill = header_fill
        ws_overall['D1'].font = header_font
        ws_overall['E1'].fill = header_fill
        ws_overall['E1'].font = header_font
        
        ws_overall['D2'] = 'Passed'
        ws_overall['E2'] = total_passed
        ws_overall['D3'] = 'Failed'
        ws_overall['E3'] = total_failed
        
        # Create pie chart
        pie = PieChart()
        pie.title = "Test Execution Results"
        pie.style = 10
        
        # Data for pie chart
        labels = Reference(ws_overall, min_col=4, min_row=2, max_row=3)
        data = Reference(ws_overall, min_col=5, min_row=1, max_row=3)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        
        # Customize data points colors
        slice_colors = ['00B050', 'FF0000']  # Green for Passed, Red for Failed
        for i, color in enumerate(slice_colors):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color
            pie.series[0].dPt.append(pt)
        
        # Position the chart next to the metrics (starting at column D, row 5)
        ws_overall.add_chart(pie, "D5")
        
        # Issues found section
        ws_overall['A7'] = 'Issues Found'
        ws_overall['A7'].font = Font(bold=True, size=12)
        ws_overall['A8'] = 'Scenario'
        ws_overall['B8'] = 'Endpoint'
        ws_overall['C8'] = 'Field'
        ws_overall['D8'] = 'Expected'
        ws_overall['E8'] = 'Actual'
        ws_overall['F8'] = 'Status'
        ws_overall['G8'] = 'Root Cause'
        ws_overall['H8'] = 'Execution Date'
        ws_overall['I8'] = 'Failed Record'
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws_overall[f'{col}8'].fill = header_fill
            ws_overall[f'{col}8'].font = header_font
        
        # Get all failed validation results
        failed_results = db.query(ValidationResult).filter(ValidationResult.status == 'fail').all()
        row = 9
        for result in failed_results:
            scenario = db.query(TestScenario).filter(TestScenario.id == result.scenario_id).first()
            execution = db.query(TestExecution).filter(TestExecution.id == result.execution_id).first()
            endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == scenario.endpoint_id).first() if scenario else None
            
            ws_overall[f'A{row}'] = scenario.name if scenario else 'Unknown'
            ws_overall[f'B{row}'] = endpoint.name if endpoint else 'Unknown'
            ws_overall[f'C{row}'] = result.field_name
            ws_overall[f'D{row}'] = result.expected or ''
            ws_overall[f'E{row}'] = result.actual or ''
            ws_overall[f'F{row}'] = result.status
            ws_overall[f'G{row}'] = result.root_cause or ''
            ws_overall[f'H{row}'] = execution.execution_date.strftime('%Y-%m-%d %H:%M:%S') if execution else ''
            
            # Add failed record data for business rule scenarios
            if scenario and scenario.category == 'business_rule' and execution and execution.response_body:
                try:
                    import json
                    # Get business rule validations for this scenario
                    business_validations = db.query(BusinessRuleValidation).filter(
                        BusinessRuleValidation.endpoint_id == scenario.endpoint_id,
                        BusinessRuleValidation.result == 'FAIL',
                        BusinessRuleValidation.field_name == result.field_name
                    ).order_by(BusinessRuleValidation.execution_date.desc()).first()
                    
                    if business_validations and business_validations.record_number:
                        # Parse response to extract records
                        response_data = json.loads(execution.response_body) if isinstance(execution.response_body, str) else execution.response_body
                        
                        # Extract array of records
                        records = []
                        if isinstance(response_data, list):
                            records = response_data
                        elif isinstance(response_data, dict):
                            for key in ['termsAndConditions', 'data', 'results', 'items', 'records']:
                                if key in response_data and isinstance(response_data[key], list):
                                    records = response_data[key]
                                    break
                        
                        # Get specific failing record
                        if records and business_validations.record_number > 0 and business_validations.record_number <= len(records):
                            specific_record = records[business_validations.record_number - 1]
                            ws_overall[f'I{row}'] = json.dumps(specific_record, indent=2)
                        else:
                            ws_overall[f'I{row}'] = 'N/A'
                    else:
                        ws_overall[f'I{row}'] = 'N/A'
                except Exception as e:
                    ws_overall[f'I{row}'] = f'Error: {str(e)}'
            else:
                ws_overall[f'I{row}'] = 'N/A'
            
            row += 1
        
        # Auto-size columns
        for col in ws_overall.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_overall.column_dimensions[column].width = adjusted_width
        
        # Sheet 2+: Per-Endpoint Results
        endpoints = db.query(APIEndpoint).all()
        
        for endpoint in endpoints:
            # Get all scenarios for this endpoint
            scenarios = db.query(TestScenario).filter(TestScenario.endpoint_id == endpoint.id).all()
            scenario_ids = [s.id for s in scenarios]
            
            if not scenario_ids:
                continue
            
            # Create sheet for this endpoint
            sheet_name = endpoint.name[:31] if len(endpoint.name) <= 31 else endpoint.name[:28] + '...'
            ws_endpoint = wb.create_sheet(sheet_name)
            
            # Endpoint info
            ws_endpoint['A1'] = 'Endpoint Name'
            ws_endpoint['B1'] = endpoint.name
            ws_endpoint['A2'] = 'Method'
            ws_endpoint['B2'] = endpoint.method
            ws_endpoint['A3'] = 'Path'
            ws_endpoint['B3'] = endpoint.path
            
            # Test results header
            ws_endpoint['A5'] = 'Scenario Name'
            ws_endpoint['B5'] = 'Description'
            ws_endpoint['C5'] = 'Status'
            ws_endpoint['D5'] = 'Pass Count'
            ws_endpoint['E5'] = 'Fail Count'
            ws_endpoint['F5'] = 'Response Time (ms)'
            ws_endpoint['G5'] = 'Execution Date'
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws_endpoint[f'{col}5'].fill = header_fill
                ws_endpoint[f'{col}5'].font = header_font
            
            # Get executions for this endpoint
            endpoint_executions = db.query(TestExecution).filter(
                TestExecution.scenario_id.in_(scenario_ids)
            ).order_by(TestExecution.execution_date.desc()).all()
            
            row = 6
            for execution in endpoint_executions:
                scenario = db.query(TestScenario).filter(TestScenario.id == execution.scenario_id).first()
                
                ws_endpoint[f'A{row}'] = scenario.name if scenario else 'Unknown'
                ws_endpoint[f'B{row}'] = scenario.description if scenario else ''
                ws_endpoint[f'C{row}'] = 'PASS' if execution.fail_count == 0 else 'FAIL'
                ws_endpoint[f'D{row}'] = execution.pass_count
                ws_endpoint[f'E{row}'] = execution.fail_count
                ws_endpoint[f'F{row}'] = execution.total_response_time_ms or 0
                ws_endpoint[f'G{row}'] = execution.execution_date.strftime('%Y-%m-%d %H:%M:%S')
                
                # Color code status
                if execution.fail_count == 0:
                    ws_endpoint[f'C{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    ws_endpoint[f'C{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                row += 1
            
            # Add failed validations for this endpoint
            ws_endpoint[f'A{row+1}'] = 'Failed Validations'
            ws_endpoint[f'A{row+1}'].font = Font(bold=True, size=12)
            
            ws_endpoint[f'A{row+2}'] = 'Scenario'
            ws_endpoint[f'B{row+2}'] = 'Field'
            ws_endpoint[f'C{row+2}'] = 'Expected'
            ws_endpoint[f'D{row+2}'] = 'Actual'
            ws_endpoint[f'E{row+2}'] = 'Root Cause'
            ws_endpoint[f'F{row+2}'] = 'Suggested Action'
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws_endpoint[f'{col}{row+2}'].fill = header_fill
                ws_endpoint[f'{col}{row+2}'].font = header_font
            
            # Get failed validations for this endpoint
            failed_validations = db.query(ValidationResult).filter(
                ValidationResult.scenario_id.in_(scenario_ids),
                ValidationResult.status == 'fail'
            ).all()
            
            detail_row = row + 3
            for validation in failed_validations:
                scenario = db.query(TestScenario).filter(TestScenario.id == validation.scenario_id).first()
                ws_endpoint[f'A{detail_row}'] = scenario.name if scenario else 'Unknown'
                ws_endpoint[f'B{detail_row}'] = validation.field_name
                ws_endpoint[f'C{detail_row}'] = validation.expected or ''
                ws_endpoint[f'D{detail_row}'] = validation.actual or ''
                ws_endpoint[f'E{detail_row}'] = validation.root_cause or ''
                ws_endpoint[f'F{detail_row}'] = validation.suggested_action or ''
                detail_row += 1
            
            # Auto-size columns
            for col in ws_endpoint.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_endpoint.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as downloadable file
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=test_execution_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
        
    finally:
        db.close()

@app.post("/api/business-rules", response_model=CustomBusinessRuleResponse)
def create_business_rule(rule: CustomBusinessRuleRequest, db: Session = Depends(get_db)):
    """Create a new custom business rule for an endpoint"""
    new_rule = CustomBusinessRule(
        endpoint_id=rule.endpoint_id,
        rule_name=rule.rule_name,
        rule_text=rule.rule_text
    )
    db.add(new_rule)
    db.commit()
    db.refresh(new_rule)
    
    # Automatically create a test scenario for this business rule
    try:
        print(f"Attempting to create test scenario for business rule: {rule.rule_name}")
        print(f"  endpoint_id: {rule.endpoint_id}")
        print(f"  business_rule_id: {new_rule.id}")
        
        test_scenario = TestScenario(
            mapping_id=None,  # Business rules don't require mapping
            endpoint_id=rule.endpoint_id,
            name=f"Business Rule: {rule.rule_name}",
            description=rule.rule_text,
            category="business_rule",
            status="pending",
            business_rule_id=new_rule.id
        )
        db.add(test_scenario)
        db.commit()
        db.refresh(test_scenario)
        print(f"✓ SUCCESS: Created test scenario with ID: {test_scenario.id} for business rule: {rule.rule_name}")
    except Exception as e:
        import traceback
        print(f"✗ ERROR creating test scenario: {str(e)}")
        print(f"Full traceback:")
        traceback.print_exc()
        # Don't fail the whole request if scenario creation fails
        db.rollback()
    
    # Extract data before session closes
    rule_data = {
        "id": new_rule.id,
        "endpoint_id": new_rule.endpoint_id,
        "rule_name": new_rule.rule_name,
        "rule_text": new_rule.rule_text,
        "created_date": new_rule.created_date,
        "updated_date": new_rule.updated_date,
        "is_active": new_rule.is_active
    }
    
    return rule_data

@app.get("/api/business-rules/endpoint/{endpoint_id}", response_model=List[CustomBusinessRuleResponse])
def get_business_rules_by_endpoint(endpoint_id: int, db: Session = Depends(get_db)):
    """Get all business rules for a specific endpoint"""
    try:
        rules = db.query(CustomBusinessRule).filter(
            CustomBusinessRule.endpoint_id == endpoint_id,
            CustomBusinessRule.is_active == True
        ).all()
        return rules
    finally:
        db.close()

@app.delete("/api/business-rules/{rule_id}")
def delete_business_rule(rule_id: int, db: Session = Depends(get_db)):
    """Delete a business rule"""
    try:
        rule = db.query(CustomBusinessRule).filter(CustomBusinessRule.id == rule_id).first()
        if not rule:
            raise HTTPException(status_code=404, detail="Rule not found")
        
        # Delete associated test scenario
        test_scenario = db.query(TestScenario).filter(TestScenario.business_rule_id == rule_id).first()
        if test_scenario:
            db.delete(test_scenario)
        
        db.delete(rule)
        db.commit()
        return {"message": "Rule deleted successfully"}
    finally:
        db.close()

@app.put("/api/business-rules/{rule_id}", response_model=CustomBusinessRuleResponse)
def update_business_rule(rule_id: int, rule_name: str = None, rule_text: str = None, db: Session = Depends(get_db)):
    """Update a business rule"""
    try:
        rule = db.query(CustomBusinessRule).filter(CustomBusinessRule.id == rule_id).first()
        if not rule:
            raise HTTPException(status_code=404, detail="Rule not found")
        
        if rule_name:
            rule.rule_name = rule_name
        if rule_text:
            rule.rule_text = rule_text
        rule.updated_date = datetime.utcnow()
        
        db.commit()
        db.refresh(rule)
        return rule
    finally:
        db.close()

def interpret_rule_with_ai(rule_text: str, schema_fields: List[str]) -> dict:
    """Use AI to interpret natural language rule into validation logic"""
    client = AzureOpenAI(
        api_key=os.getenv("AZURE_OPENAI_API_KEY"),
        api_version=os.getenv("AZURE_OPENAI_API_VERSION"),
        azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT")
    )
    deployment_name = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME")
    
    prompt = f"""You are a validation rule interpreter. Convert the following natural language business rule into a structured validation format.

Available fields in the API response: {', '.join(schema_fields)}

Business Rule: "{rule_text}"

Return a JSON object with this structure:
{{
    "condition": "the IF condition or null if unconditional",
    "condition_field": "field name in condition or null",
    "condition_operator": "operator like 'startswith', 'equals', 'contains', 'is_null', 'is_not_null', 'greater_than', 'less_than'",
    "condition_value": "value to check in condition or null",
    "validation_field": "field to validate",
    "validation_operator": "operator like 'must_be', 'must_not_be', 'must_be_one_of', 'must_be_null', 'must_not_be_null', 'must_be_before', 'must_be_after'",
    "validation_value": "expected value or list of values",
    "error_message": "user-friendly error message"
}}

Examples:
Rule: "If title starts with HRSA- then hrsaWideFlag must be Y"
Output: {{"condition": "title starts with HRSA-", "condition_field": "title", "condition_operator": "startswith", "condition_value": "HRSA-", "validation_field": "hrsaWideFlag", "validation_operator": "must_be", "validation_value": "Y", "error_message": "hrsaWideFlag must be Y when title starts with HRSA-"}}

Rule: "trackingTypeCode must be one of: Term, Condition, Reporting Requirement"
Output: {{"condition": null, "condition_field": null, "condition_operator": null, "condition_value": null, "validation_field": "trackingTypeCode", "validation_operator": "must_be_one_of", "validation_value": ["Term", "Condition", "Reporting Requirement"], "error_message": "trackingTypeCode must be one of: Term, Condition, Reporting Requirement"}}

Rule: "activeDate should not be later than lastUpdateDate"
Output: {{"condition": null, "condition_field": null, "condition_operator": null, "condition_value": null, "validation_field": "activeDate", "validation_operator": "must_be_before", "validation_value": "lastUpdateDate", "error_message": "activeDate should not be later than lastUpdateDate"}}

Rule: "If programType is null then inactiveDate must also be null"
Output: {{"condition": "programType is null", "condition_field": "programType", "condition_operator": "is_null", "condition_value": null, "validation_field": "inactiveDate", "validation_operator": "must_be_null", "validation_value": null, "error_message": "inactiveDate must be null when programType is null"}}

Return ONLY the JSON object, no additional text."""

    try:
        response = client.chat.completions.create(
            model=deployment_name,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=500,
            temperature=0.1
        )
        
        response_text = response.choices[0].message.content.strip()
        
        # Extract JSON from response
        start_idx = response_text.find('{')
        end_idx = response_text.rfind('}') + 1
        if start_idx != -1 and end_idx > start_idx:
            json_str = response_text[start_idx:end_idx]
            result = json.loads(json_str)
            return result
        else:
            raise ValueError("No JSON found in AI response")
            
    except Exception as e:
        print(f"Error interpreting rule with AI: {e}")
        return {
            "condition": None,
            "condition_field": None,
            "condition_operator": None,
            "condition_value": None,
            "validation_field": None,
            "validation_operator": None,
            "validation_value": None,
            "error_message": f"Failed to interpret rule: {str(e)}"
        }

def apply_validation_rule(record: dict, rule_interpretation: dict, record_number: int) -> dict:
    """Apply interpreted validation rule to a single record"""
    
    # Check if condition is met (if there is one)
    condition_met = True
    if rule_interpretation.get("condition_field"):
        field_value = record.get(rule_interpretation["condition_field"])
        operator = rule_interpretation.get("condition_operator")
        condition_value = rule_interpretation.get("condition_value")
        
        if operator == "startswith":
            condition_met = str(field_value or "").startswith(str(condition_value))
        elif operator == "equals":
            condition_met = str(field_value) == str(condition_value)
        elif operator == "contains":
            condition_met = str(condition_value) in str(field_value or "")
        elif operator == "is_null":
            condition_met = field_value is None or field_value == "" or field_value == "null"
        elif operator == "is_not_null":
            condition_met = field_value is not None and field_value != "" and field_value != "null"
        elif operator == "greater_than":
            try:
                condition_met = float(field_value) > float(condition_value)
            except:
                condition_met = False
        elif operator == "less_than":
            try:
                condition_met = float(field_value) < float(condition_value)
            except:
                condition_met = False
    
    # If condition not met, validation passes (rule doesn't apply)
    if not condition_met:
        return {
            "record_number": record_number,
            "result": "PASS",
            "field_name": rule_interpretation.get("validation_field"),
            "expected_value": None,
            "actual_value": None,
            "error_message": "Condition not met, rule does not apply"
        }
    
    # Apply validation
    validation_field = rule_interpretation.get("validation_field")
    actual_value = record.get(validation_field)
    validation_operator = rule_interpretation.get("validation_operator")
    validation_value = rule_interpretation.get("validation_value")
    
    result = "PASS"
    error_message = None
    
    if validation_operator == "must_be":
        if str(actual_value) != str(validation_value):
            result = "FAIL"
            error_message = rule_interpretation.get("error_message", f"{validation_field} must be {validation_value}")
    
    elif validation_operator == "must_not_be":
        if str(actual_value) == str(validation_value):
            result = "FAIL"
            error_message = rule_interpretation.get("error_message", f"{validation_field} must not be {validation_value}")
    
    elif validation_operator == "must_be_one_of":
        if str(actual_value) not in [str(v) for v in validation_value]:
            result = "FAIL"
            error_message = rule_interpretation.get("error_message", f"{validation_field} must be one of {validation_value}")
    
    elif validation_operator == "must_be_null":
        if actual_value is not None and actual_value != "" and actual_value != "null":
            result = "FAIL"
            error_message = rule_interpretation.get("error_message", f"{validation_field} must be null")
    
    elif validation_operator == "must_not_be_null":
        if actual_value is None or actual_value == "" or actual_value == "null":
            result = "FAIL"
            error_message = rule_interpretation.get("error_message", f"{validation_field} must not be null")
    
    elif validation_operator == "must_be_before":
        # Compare dates
        try:
            from dateutil import parser
            actual_date = parser.parse(str(actual_value))
            
            # Check if validation_value is a field name or a date string
            if validation_value in record:
                expected_date = parser.parse(str(record[validation_value]))
            else:
                expected_date = parser.parse(str(validation_value))
            
            if actual_date > expected_date:
                result = "FAIL"
                error_message = rule_interpretation.get("error_message", f"{validation_field} must be before {validation_value}")
        except:
            result = "FAIL"
            error_message = f"Could not parse dates for comparison"
    
    elif validation_operator == "must_be_after":
        # Compare dates
        try:
            from dateutil import parser
            actual_date = parser.parse(str(actual_value))
            
            # Check if validation_value is a field name or a date string
            if validation_value in record:
                expected_date = parser.parse(str(record[validation_value]))
            else:
                expected_date = parser.parse(str(validation_value))
            
            if actual_date < expected_date:
                result = "FAIL"
                error_message = rule_interpretation.get("error_message", f"{validation_field} must be after {validation_value}")
        except:
            result = "FAIL"
            error_message = f"Could not parse dates for comparison"
    
    return {
        "record_number": record_number,
        "result": result,
        "field_name": validation_field,
        "expected_value": str(validation_value) if validation_value else None,
        "actual_value": str(actual_value) if actual_value else None,
        "error_message": error_message
    }

@app.post("/api/business-rules/validate/{endpoint_id}")
async def run_business_rule_validation(endpoint_id: int, db: Session = Depends(get_db)):
    """Run all business rules for an endpoint against its API response"""
    try:
        # Get endpoint
        endpoint = db.query(APIEndpoint).filter(APIEndpoint.id == endpoint_id).first()
        if not endpoint:
            raise HTTPException(status_code=404, detail="Endpoint not found")
        
        # Get all active rules for this endpoint
        rules = db.query(CustomBusinessRule).filter(
            CustomBusinessRule.endpoint_id == endpoint_id,
            CustomBusinessRule.is_active == True
        ).all()
        
        if not rules:
            return {"message": "No active rules found for this endpoint", "results": []}
        
        # Call the API endpoint
        url = endpoint.base_url + endpoint.path
        headers = json.loads(endpoint.headers) if endpoint.headers else {}
        
        # Handle authentication
        if endpoint.auth_type == "bearer" and endpoint.auth_token:
            headers["Authorization"] = f"Bearer {endpoint.auth_token}"
        
        # Make API call
        if endpoint.method.upper() == "GET":
            response = requests.get(url, headers=headers, timeout=endpoint.timeout_ms/1000)
        elif endpoint.method.upper() == "POST":
            body = json.loads(endpoint.default_request_body) if endpoint.default_request_body else {}
            response = requests.post(url, json=body, headers=headers, timeout=endpoint.timeout_ms/1000)
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported method: {endpoint.method}")
        
        if response.status_code != 200:
            raise HTTPException(status_code=response.status_code, detail=f"API call failed: {response.text}")
        
        # Parse response
        response_data = response.json()
        
        # Determine if response is a list or single object
        records = []
        wrapper_fields = {}
        array_key = None
        
        if isinstance(response_data, list):
            records = response_data
        elif isinstance(response_data, dict):
            # Look for arrays in the response object
            # Common patterns: "data", "items", "results", or any array property
            found_array = False
            for key, value in response_data.items():
                if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                    # Found an array of objects - use this as records
                    records = value
                    array_key = key
                    found_array = True
                    # Store wrapper-level fields (everything except the array)
                    wrapper_fields = {k: v for k, v in response_data.items() if k != key}
                    break
            
            # If no array found, treat the whole object as a single record
            if not found_array:
                records = [response_data]
        else:
            records = [response_data]
        
        # Get schema fields from both wrapper and first record
        # Include wrapper fields (totalRecords, limit, etc.) and record fields (title, hrsaWideFlag, etc.)
        schema_fields = []
        if wrapper_fields:
            # Add wrapper-level fields
            for key in wrapper_fields.keys():
                if not isinstance(wrapper_fields[key], dict):  # Skip nested objects like "links"
                    schema_fields.append(key)
        
        # Add record-level fields
        if records and len(records) > 0:
            schema_fields.extend(list(records[0].keys()))
        
        # Run each rule against all records
        all_results = []
        
        for rule in rules:
            # Interpret the rule using AI
            rule_interpretation = interpret_rule_with_ai(rule.rule_text, schema_fields)
            
            # Check if referenced fields exist in schema
            validation_field = rule_interpretation.get("validation_field")
            condition_field = rule_interpretation.get("condition_field")
            
            field_warnings = []
            if validation_field and validation_field not in schema_fields:
                field_warnings.append(f"Warning: Field '{validation_field}' not found in API response schema")
            if condition_field and condition_field not in schema_fields:
                field_warnings.append(f"Warning: Field '{condition_field}' not found in API response schema")
            
            # Determine if this rule references wrapper-level fields or array-level fields
            wrapper_field_keys = list(wrapper_fields.keys()) if wrapper_fields else []
            is_wrapper_rule = (validation_field in wrapper_field_keys) or (condition_field in wrapper_field_keys)
            
            if is_wrapper_rule:
                # Validate wrapper-level fields once (not per record)
                # Merge wrapper fields with first record for validation context
                wrapper_record = {**wrapper_fields, **(records[0] if records else {})}
                validation_result = apply_validation_rule(wrapper_record, rule_interpretation, 0)
                validation_result["record_number"] = None  # Wrapper validation, not tied to a specific record
                
                # Save validation result to database
                db_result = BusinessRuleValidation(
                    rule_id=rule.id,
                    endpoint_id=endpoint_id,
                    record_number=validation_result["record_number"],
                    field_name=validation_result["field_name"],
                    expected_value=validation_result["expected_value"],
                    actual_value=validation_result["actual_value"],
                    result=validation_result["result"],
                    error_message=validation_result["error_message"],
                    rule_text=rule.rule_text
                )
                db.add(db_result)
                
                all_results.append({
                    "rule_id": rule.id,
                    "rule_name": rule.rule_name,
                    "rule_text": rule.rule_text,
                    "record_number": "Wrapper",
                    "field_name": validation_result["field_name"],
                    "expected_value": validation_result["expected_value"],
                    "actual_value": validation_result["actual_value"],
                    "result": validation_result["result"],
                    "error_message": validation_result["error_message"],
                    "warnings": field_warnings
                })
            else:
                # Apply rule to each record in the array
                for idx, record in enumerate(records, start=1):
                    # Merge wrapper fields into each record for context
                    enriched_record = {**wrapper_fields, **record}
                    validation_result = apply_validation_rule(enriched_record, rule_interpretation, idx)
                    
                    # Save validation result to database
                    db_result = BusinessRuleValidation(
                        rule_id=rule.id,
                        endpoint_id=endpoint_id,
                        record_number=validation_result["record_number"],
                        field_name=validation_result["field_name"],
                        expected_value=validation_result["expected_value"],
                        actual_value=validation_result["actual_value"],
                        result=validation_result["result"],
                        error_message=validation_result["error_message"],
                        rule_text=rule.rule_text
                    )
                    db.add(db_result)
                    
                    all_results.append({
                        "rule_id": rule.id,
                        "rule_name": rule.rule_name,
                        "rule_text": rule.rule_text,
                        "record_number": validation_result["record_number"],
                        "field_name": validation_result["field_name"],
                        "expected_value": validation_result["expected_value"],
                        "actual_value": validation_result["actual_value"],
                        "result": validation_result["result"],
                        "error_message": validation_result["error_message"],
                        "warnings": field_warnings
                    })
        
        db.commit()
        
        # Calculate summary
        total_validations = len(all_results)
        passed = sum(1 for r in all_results if r["result"] == "PASS")
        failed = sum(1 for r in all_results if r["result"] == "FAIL")
        
        return {
            "summary": {
                "total_validations": total_validations,
                "passed": passed,
                "failed": failed,
                "total_records": len(records),
                "total_rules": len(rules)
            },
            "results": all_results,
            "api_response": response_data
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Validation failed: {str(e)}")
    finally:
        db.close()

@app.get("/api/business-rules/validation-history/{endpoint_id}")
def get_validation_history(endpoint_id: int, db: Session = Depends(get_db)):
    """Get validation history for an endpoint"""
    try:
        results = db.query(BusinessRuleValidation).filter(
            BusinessRuleValidation.endpoint_id == endpoint_id
        ).order_by(BusinessRuleValidation.execution_date.desc()).limit(100).all()
        
        return results
    finally:
        db.close()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
